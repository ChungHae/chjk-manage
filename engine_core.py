# -*- coding: utf-8 -*-
"""입출금·미수 정리 엔진 (영구본). 규칙: _정리규칙_명세서.md 참조."""
import warnings, re, os, glob, calendar
from copy import copy
from datetime import date, timedelta
warnings.filterwarnings('ignore')
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

TODAY = date(2026, 6, 9)
STD_FONT = Font(name="맑은 고딕", size=9)
HDR_FONT = Font(name="맑은 고딕", size=9, bold=True)
CEN = Alignment(horizontal="center", vertical="center")
ACCT = '_-* #,##0_-;\\-* #,##0_-;_-* "-"_-;_-@_-'   # 회계 서식
DATEFMT = 'mm-dd-yy'
WRAP=Alignment(horizontal="center",vertical="center",wrap_text=True)
def apply_formats(ws, C):
    import datetime as _dt
    amt={C['공급'],C['부가'],C['합계'],C['정산'],C['입금'],C['차액']}
    dts={C['작성'],C['입금일']}
    for r in range(1,ws.max_row+1):
        wrapped=False
        for c in range(1,C['n']+1):
            cell=ws.cell(r,c)
            if isinstance(cell,MergedCell): continue
            if isinstance(cell.value,str) and "\n" in cell.value:
                cell.alignment=WRAP; wrapped=True
            else:
                cell.alignment=CEN
            if r>=3:
                if c in dts: cell.number_format=DATEFMT
                elif c in amt: cell.number_format=ACCT
        if wrapped: ws.row_dimensions[r].height=30
LEGAL = re.compile(r"주식회사|\(주\)|㈜|\(유\)|유한회사|\(주\)|（주）|主식회사")  # (자)는 표준 법인격 아님 → 파일명서 제거 안 함

# ---------- 파서 ----------
def _n(s): return re.sub(r"\s+", "", str(s)) if s is not None else ""
def _num(v):
    if v is None or (isinstance(v,float) and pd.isna(v)): return 0
    s=re.sub(r"[^\d\-.]","",str(v))
    try: return int(round(float(s))) if s not in("","-",".") else 0
    except: return 0
def _bdate(v):
    if v is None or (isinstance(v,float) and pd.isna(v)): return None
    s=str(v).strip().split("\n")[0].replace(".","-").replace("/","-")
    m=re.search(r"(\d{4})-(\d{1,2})-(\d{1,2})",s)
    if m: return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    m=re.search(r"\b(\d{4})(\d{2})(\d{2})\b",s)
    if m: return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    return None

def parse_bank(path, bank):
    raw=pd.read_excel(path,header=None,dtype=object); hr=None
    for i in range(min(15,len(raw))):
        c=[_n(x) for x in raw.iloc[i].tolist()]
        if any("입금" in x for x in c) and any("거래일" in x for x in c): hr=i; break
    if hr is None: raise ValueError(f"[{bank}] 헤더없음 {path}")
    h=[_n(x) for x in raw.iloc[hr].tolist()]
    def cc(must,exc=()):
        for i,x in enumerate(h):
            if must in x and not any(e in x for e in exc): return i
        return None
    NAME=["상대계좌예금주명","예금주명","기재내용","의뢰인/수취인","거래기록사항","거래내용","내용","적요"]
    def cp():
        for k in NAME:
            for i,x in enumerate(h):
                if x==_n(k): return i
        for k in NAME:
            for i,x in enumerate(h):
                if _n(k) in x: return i
        return None
    cd=cc("거래일"); cdep=cc("입금"); cnm=cp()
    # 이름후보칸 전부 수집(우선순위 순) → 행별로 첫 비지않은 값 사용(하나 '대체' 적요 등 폴백)
    namecols=[]
    for k in NAME:
        for i,x in enumerate(h):
            if x==_n(k) or _n(k) in x:
                if i not in namecols: namecols.append(i)
    if cnm is not None and cnm not in namecols: namecols.insert(0,cnm)
    out=[]
    for _,r in raw.iloc[hr+1:].iterrows():
        d=_bdate(r.iloc[cd]) if cd is not None else None
        if not d: continue
        nm=""
        for ci in namecols:
            if ci < len(r) and not pd.isna(r.iloc[ci]):
                v=str(r.iloc[ci]).strip()
                if v and v.lower()!="nan": nm=v; break
        out.append(dict(은행=bank,거래일=d,입금액=_num(r.iloc[cdep]) if cdep is not None else 0,상대방명=nm))
    return [x for x in out if x['입금액']>0]

def _parse_tax(path, who_col):
    raw=pd.read_excel(path,header=None,dtype=object); hr=None
    for i in range(min(12,len(raw))):
        if any(str(x).strip()=="작성일자" for x in raw.iloc[i].tolist()): hr=i; break
    df=pd.read_excel(path,header=hr,dtype=object)
    df=df[df['작성일자'].astype(str).str.match(r'\d{4}-\d{2}-\d{2}')].copy()
    cols=list(df.columns)
    if who_col=="공급받는자":
        cb=[c for c in cols if "공급받는자사업자등록번호" in str(c)]
        cb=cb[0] if cb else [c for c in cols if "사업자등록번호" in str(c)][1]
        si=[i for i,c in enumerate(cols) if str(c).startswith("상호")]
        cn=cols[si[1]] if len(si)>1 else cols[si[0]]
    else:
        cb=[c for c in cols if "공급자사업자등록번호" in str(c)][0]
        cn=cols[[i for i,c in enumerate(cols) if str(c).startswith("상호")][0]]
    out=[]
    for _,r in df.iterrows():
        try: amt=int(round(float(re.sub(r"[^\d\-.]","",str(r['합계금액'])))))
        except: continue
        sup=int(round(float(re.sub(r"[^\d\-.]","",str(r['공급가액']))))) if not pd.isna(r['공급가액']) else None
        out.append(dict(작성일=str(r['작성일자'])[:10],승인번호=str(r['승인번호']).strip(),
            거래처=re.sub(r"\s+"," ",str(r[cn]).replace("（","(").replace("）",")")).strip(),
            사업자번호=str(r[cb]).strip(),공급가=sup,합계=amt))
    return out
def parse_sales(path): return _parse_tax(path,"공급받는자")
def parse_purchase(path): return _parse_tax(path,"공급자")

def parse_eum(path, bank="신한"):
    """어음/판매채권 수취내역 → 발행인·금액·수취일·만기일. 신한(수취내역)·하나(판매채권내역조회) 형식 모두 지원."""
    raw=pd.read_excel(path,header=None,dtype=object)
    hr=None
    for i in range(min(15,len(raw))):
        row=" ".join(str(x) for x in raw.iloc[i].tolist())
        if ("만기" in row) and (("발행인" in row) or ("구매기업" in row)): hr=i; break
    df=pd.read_excel(path,header=(hr if hr is not None else 0),dtype=object)
    cols=list(df.columns)
    def find(*keys):
        for k in keys:
            for c in cols:
                if k in str(c): return c
        return None
    c_name=find("발행인","구매기업"); c_amt=find("어음금액","채권금액","금액")
    c_recv=find("수취일","발행일"); c_mat=find("만기일","만기"); c_stat=find("상태")
    out=[]
    if c_name is None: return out
    for _,r in df.iterrows():
        fa=r.get(c_name)
        if fa is None or (isinstance(fa,float) and pd.isna(fa)): continue
        if str(fa).strip() in ('','합계','총계','소계','nan'): continue
        amt=_num(r.get(c_amt)); recv=_bdate(r.get(c_recv)); mat=_bdate(r.get(c_mat))
        if not recv or amt<=0: continue
        out.append(dict(발행인=str(fa).strip(),
                        어음금액=amt, 수취일=recv, 만기일=mat,
                        상태=str(r.get(c_stat,'') if c_stat else '').strip(), bank=f"{bank} (어음)"))
    return out

# ---------- 매칭 (퍼지 금지) ----------
def cname(s):
    s=str(s).replace("（","(").replace("）",")")
    s=re.sub(r"\(주\)|주식회사|㈜|\(유\)|유한회사|\(자\)|영업소|前","",s)
    s=re.sub(r"\([^)]*\)","",s)
    return re.sub(r"[^가-힣A-Za-z0-9]","",s).lower()
NONAR=re.compile(r"지방소득|환급|국세|부가가치|이자|결제|결:|카드|페이|페이먼츠|BC|KB\d|삼성\d|신한\d|하나\d|국민\d|우리\d|^\d|수수료|급여|보험|공단|렌탈|리스|쿠팡|네이버|배달|토스|kakao|KIA|캐피탈|SMS|통지|세무서|자동차세|해상|스토아|캐쉬백|관세|Npay")

def match_deposit(loc, dep_name, universe, alias):
    """returns (kind, 거래처명 or None). 퍼지 매칭 없음."""
    key=(loc, cname(dep_name))
    if key in alias:
        return ('제외', None) if str(alias[key]).strip()=='제외' else ('별칭', alias[key])
    key2=(loc, re.sub(r'\d+$','',cname(dep_name)))   # 숫자접미 제거(예: sk상사0508→sk상사) 별칭
    if key2!=key and key2[1] and key2 in alias:
        return ('제외', None) if str(alias[key2]).strip()=='제외' else ('별칭', alias[key2])
    if NONAR.search(str(dep_name)): return ('제외', None)
    nd=cname(dep_name); paren=re.findall(r"[(（]([^)）]+)",str(dep_name))
    hints=[nd]+[cname(p) for p in paren if cname(p)]
    for h in hints:
        if not h: continue
        c=list(set(v for k,v in universe.get(loc,{}).items()
                   if k==h or (len(h)>=2 and (k.startswith(h) or h.startswith(k)))))
        if len(c)==1: return ('자동', c[0])
        if len(c)>1: return ('다중', None)
    return ('미배정', None)

# ---------- 스타일/너비 ----------
def disp_len(v):
    import datetime as _dt
    if isinstance(v,(_dt.datetime,_dt.date)): return 9   # 내장 간단한날짜(지역설정 연-월-일)
    if isinstance(v,(int,float)): return len(f"{v:,.0f}")
    s="" if v is None else str(v)
    if "\n" in s:   # 어음 등 줄바꿈 → 줄별 최대폭
        return max(disp_len(line) for line in s.split("\n"))
    w=0
    for ch in s:
        w += 2 if ord(ch)>0x1100 else 1
    return w
def autofit(ws, ncol):
    # 수식(금액) 칸은 계산값을 못 재므로 시트 내 최대 숫자폭(회계여백 포함) 기준으로
    maxnum=11
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell,MergedCell): continue
            if isinstance(cell.value,(int,float)): maxnum=max(maxnum, len(f"{cell.value:,.0f}")+3)
    for c in range(1,ncol+1):
        mx=8
        for r in range(1,ws.max_row+1):
            cell=ws.cell(r,c)
            if isinstance(cell,MergedCell): continue
            v=cell.value
            if isinstance(v,str) and v.startswith("="):
                mx=max(mx,maxnum); continue
            mx=max(mx, disp_len(v)+2)
        ws.column_dimensions[get_column_letter(c)].width=min(max(mx,8),60)
def _fi(ws, C, r):
    """행 r의 정산액 F(정산수식이 참조하는 합계 합)와 입금합 I 반환. 차액수식 없으면 None."""
    jf=ws.cell(r,C['차액']).value
    if not (isinstance(jf,str) and jf.startswith("=")): return None
    HE=get_column_letter(C['합계']); IE=get_column_letter(C['입금'])
    sf=ws.cell(r,C['정산']).value; rows_h=[]
    if isinstance(sf,str):
        for m in re.finditer(r"SUM\("+HE+r"(\d+):"+HE+r"(\d+)\)", sf):
            rows_h += list(range(int(m.group(1)),int(m.group(2))+1))
        sf2=re.sub(r"SUM\("+HE+r"\d+:"+HE+r"\d+\)","",sf)
        for m in re.finditer(HE+r"(\d+)", sf2): rows_h.append(int(m.group(1)))
    if not rows_h: rows_h=[r]
    F=sum(round((ws.cell(rh,C['공급']).value or 0)*1.1) for rh in rows_h if isinstance(ws.cell(rh,C['공급']).value,(int,float)))
    rows_i=set(int(m.group(1)) for m in re.finditer(IE+r"(\d+)", jf))
    for rh in rows_h:
        if isinstance(ws.cell(rh,C['입금']).value,(int,float)): rows_i.add(rh)
    I=sum((ws.cell(ri,C['입금']).value or 0) for ri in rows_i if isinstance(ws.cell(ri,C['입금']).value,(int,float)))
    return F, I

def _rd(x, dig):
    """ROUNDDOWN(x, dig) — Excel 방식(0 방향 절사)."""
    if dig>=0: return x
    f=10**(-dig)
    return (int(x)//f)*f if x>=0 else -(((-int(x))//f)*f)

def apply_jeolsang(ws, C):
    """차액 ROUNDDOWN 규칙:
    - 정확매칭(raw 차액=0): ROUNDDOWN 제거 → =F-I.
    - 의도적 write-off(기존 ROUNDDOWN이 raw를 0으로 만드는 경우, 비고 '절사' 등): 그대로 보존.
    - 진짜 미수/부분입금(|raw|>1000, write-off 아님): ROUNDDOWN 없이 전액 표시(=F-I).
    - ≤2025 & 0<|raw|<=1000(소액 잔차): 1000원 절사(ROUNDDOWN -3), 없으면 추가.
    - 2026~: ROUNDDOWN 없이 =F-I."""
    FE_=get_column_letter(C['정산']); IE_=get_column_letter(C['입금'])
    for r in range(3, ws.max_row+1):
        jc=ws.cell(r,C['차액'])
        if isinstance(jc,MergedCell) or not (isinstance(jc.value,str) and jc.value.startswith("=")): continue
        # 자동 생성 패턴(=F{r}-I{k}... 또는 ROUNDDOWN 래핑)만 관리. 그 외(사용자 수동 상계 등)는 보존.
        _inner=re.sub(r"^=ROUNDDOWN\((.*),\s*-?\d+\)\s*$", r"=\1", jc.value)
        if re.fullmatch(rf"={FE_}{r}(-{IE_}\d+)*", _inner) is None: continue
        a=pdate(ws.cell(r,C['작성']).value)
        fi=_fi(ws,C,r)
        if fi is None: continue
        raw=fi[0]-fi[1]
        has_rd="ROUNDDOWN" in jc.value
        dm=re.search(r",\s*(-?\d+)\)", jc.value); dig=int(dm.group(1)) if dm else None
        strip=lambda v: re.sub(r"^=ROUNDDOWN\((.*),\s*-?\d+\)\s*$", r"=\1", v)
        if raw==0:
            if has_rd: jc.value=strip(jc.value)
        elif has_rd and dig is not None and _rd(raw,dig)==0:
            pass  # 의도적 write-off → 보존
        elif abs(raw)>1000:
            if has_rd: jc.value=strip(jc.value)   # 진짜 미수: 전액 표시
        elif a and a.year<=2025:
            if not has_rd: jc.value="=ROUNDDOWN("+jc.value[1:]+",-3)"
        else:
            if has_rd: jc.value=strip(jc.value)

def apply_font(ws):
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c,MergedCell): continue
            c.font = HDR_FONT if (c.row<=2 and c.value not in (None,"")) else STD_FONT

# ---------- 컬럼/날짜 ----------
def _has_issuer(ws): return str(ws.cell(2,2).value).strip()=="발행처"
def cols(ws):
    if _has_issuer(ws): return dict(작성=1,발행=2,공급=3,부가=4,합계=5,정산=6,은행=7,입금일=8,입금=9,차액=10,비고=11,n=11)
    return dict(작성=1,발행=None,공급=2,부가=3,합계=4,정산=5,은행=6,입금일=7,입금=8,차액=9,비고=10,n=10)
def pdate(v):
    if v is None: return None
    m=re.search(r"(\d{4})[-./](\d{1,2})[-./](\d{1,2})",str(v).split("\n")[0])
    if not m: return None
    try: return date(int(m.group(1)),int(m.group(2)),int(m.group(3)))
    except: return None
def last_data_row(ws,C):
    for r in range(ws.max_row,2,-1):
        if any(ws.cell(r,C[k]).value not in (None,"") for k in ('작성','입금일','공급','입금') if C[k]):
            return r
    return 2
def due_of(d,rule="익월말"):
    # 숫자 일자: (당/익/익익/익익익)월{N}일
    md=re.match(r"(당|익익익|익익|익)월(\d+)일$", str(rule).replace(" ",""))
    if md:
        n={"당":0,"익":1,"익익":2,"익익익":3}[md.group(1)]; idx=(d.month-1)+n
        y=d.year+idx//12; mo=idx%12+1; day=min(int(md.group(2)), calendar.monthrange(y,mo)[1])
        return date(y,mo,day)
    # 일반 파싱: (당/익/익익/익익익)월(초=10일/중순=15일/말=말일)
    mm=re.match(r"(당|익익익|익익|익)월(초|중순|중|말)", str(rule).replace(" ",""))
    if mm:
        n={"당":0,"익":1,"익익":2,"익익익":3}[mm.group(1)]; part=mm.group(2)
        idx=(d.month-1)+n; y=d.year+idx//12; mo=idx%12+1
        if part=="초": return date(y,mo,10)
        if part in ("중","중순"): return date(y,mo,15)
        return date(y,mo,calendar.monthrange(y,mo)[1])
    n={"당월말":0,"익월말":1,"익익월말":2,"익익익월말":3}.get(rule,1); idx=(d.month-1)+n
    return date(d.year+idx//12, idx%12+1, calendar.monthrange(d.year+idx//12,idx%12+1)[1])
def status(d,today=TODAY,rule="익월말"):
    due=due_of(d,rule); idx=due.month; y=due.year+idx//12; m=idx%12+1
    longd=date(y,m,calendar.monthrange(y,m)[1])
    if today<=due: return "기한내"
    return "미수" if today<=longd else "장기미수"

def resolve_rule(rule, sup):
    """조건부 기준 '공급>{금액}?{참}:{거짓}' 을 계산서 공급가로 풀어 실제 기준 문자열 반환."""
    m=re.match(r"공급>(\d+)\?(.+):(.+)$", str(rule).replace(" ",""))
    if m: return m.group(2) if (sup or 0) > int(m.group(1)) else m.group(3)
    return rule

def compute_unpaid(ws, C, exc=None):
    """계산서별 미수액 = 정산액(F) - 입금합(I). 정산/차액 수식을 파싱해 그룹 단위로 계산.
    반환: [(작성일, 미수액)]. 예외(작성일,합계) 행은 제외. 단일 워크북(수식)만 필요."""
    exc=exc or set()
    HE=get_column_letter(C['합계']); IE=get_column_letter(C['입금'])
    out=[]
    for r in range(3, ws.max_row+1):
        jf=ws.cell(r,C['차액']).value
        if not (isinstance(jf,str) and jf.startswith("=")): continue
        a=pdate(ws.cell(r,C['작성']).value)
        if not a: continue
        sup0=ws.cell(r,C['공급']).value
        if isinstance(sup0,(int,float)) and ((a.isoformat(),round(sup0*1.1)) in exc or (a.isoformat(),int(sup0)) in exc): continue
        sf=ws.cell(r,C['정산']).value; rows_h=[]
        if isinstance(sf,str):
            for m in re.finditer(r"SUM\("+HE+r"(\d+):"+HE+r"(\d+)\)", sf):
                rows_h += list(range(int(m.group(1)),int(m.group(2))+1))
            sf2=re.sub(r"SUM\("+HE+r"\d+:"+HE+r"\d+\)","",sf)
            for m in re.finditer(HE+r"(\d+)", sf2): rows_h.append(int(m.group(1)))
        if not rows_h: rows_h=[r]
        F=sum(round((ws.cell(rh,C['공급']).value or 0)*1.1) for rh in rows_h if isinstance(ws.cell(rh,C['공급']).value,(int,float)))
        rows_i=set(int(m.group(1)) for m in re.finditer(IE+r"(\d+)", jf))
        # 차액수식에 입금참조가 누락된 경우(원본 오류) 보강: 정산 그룹 행의 입금도 포함
        for rh in rows_h:
            if isinstance(ws.cell(rh,C['입금']).value,(int,float)): rows_i.add(rh)
        I=sum((ws.cell(ri,C['입금']).value or 0) for ri in rows_i if isinstance(ws.cell(ri,C['입금']).value,(int,float)))
        unpaid=F-I
        mdig=re.search(r"ROUNDDOWN\([^,]*,\s*(-?\d+)\)", jf)  # 행별 절사 자릿수 반영(-3,-4,-5)
        if mdig:
            dig=int(mdig.group(1))
            if dig<0:
                factor=10**(-dig)
                unpaid=(int(unpaid)//factor)*factor if unpaid>=0 else -(((-int(unpaid))//factor)*factor)
        out.append((a, round(unpaid)))
    return out

def loose_credit(ws, C):
    """어느 계산서 차액수식에도 참조되지 않은 floating 입금 합(=미배분 입금). 케이테크처럼 금액이 안 맞아도 오래된 미수부터 차감하는 데 사용."""
    IE=get_column_letter(C['입금']); referenced=set()
    for r in range(3, ws.max_row+1):
        jf=ws.cell(r,C['차액']).value
        if isinstance(jf,str) and jf.startswith("="):
            referenced.update(int(m.group(1)) for m in re.finditer(IE+r"(\d+)", jf))
    tot=0
    for r in range(3, ws.max_row+1):
        if r in referenced: continue
        if pdate(ws.cell(r,C['작성']).value): continue   # 계산서 행은 제외 — 작성일 없는 순수 floating 입금만
        dm=ws.cell(r,C['입금']).value
        if isinstance(dm,(int,float)) and dm>0: tot+=dm
    return tot

def _net_unpaid_rows(ws, C, exc=None):
    """계산서별 net 미수(과입금·미배분 크레딧 반영) → [(row, date, net>1000)] 오래된 순."""
    HE=get_column_letter(C['합계']); IE=get_column_letter(C['입금']); per=[]
    for r in range(3, ws.max_row+1):
        jf=ws.cell(r,C['차액']).value
        if not (isinstance(jf,str) and jf.startswith("=")): continue
        a=pdate(ws.cell(r,C['작성']).value)
        if not a: continue
        sf=ws.cell(r,C['정산']).value; rows_h=[]
        if isinstance(sf,str):
            for m in re.finditer(r"SUM\("+HE+r"(\d+):"+HE+r"(\d+)\)", sf):
                rows_h += list(range(int(m.group(1)),int(m.group(2))+1))
            sf2=re.sub(r"SUM\("+HE+r"\d+:"+HE+r"\d+\)","",sf)
            for m in re.finditer(HE+r"(\d+)", sf2): rows_h.append(int(m.group(1)))
        if not rows_h: rows_h=[r]
        F=sum(round((ws.cell(rh,C['공급']).value or 0)*1.1) for rh in rows_h if isinstance(ws.cell(rh,C['공급']).value,(int,float)))
        rows_i=set(int(m.group(1)) for m in re.finditer(IE+r"(\d+)", jf))
        for rh in rows_h:
            if isinstance(ws.cell(rh,C['입금']).value,(int,float)): rows_i.add(rh)
        I=sum((ws.cell(ri,C['입금']).value or 0) for ri in rows_i if isinstance(ws.cell(ri,C['입금']).value,(int,float)))
        per.append((r, a, round(F-I)))
    credit=sum(-u for _,_,u in per if u<0) + loose_credit(ws,C)
    short=sorted([[r,a,u] for r,a,u in per if u>1000], key=lambda x:x[1])
    for it in short:
        use=min(it[2],credit); it[2]-=use; credit-=use
    return [(r,a,u) for r,a,u in short if u>1000]

def _fifo_place(ws, C, leftover, EL):
    """미배분 입금을 가장 오래된 net 미수 계산서 행에 직접 얹는다(시각적 FIFO). 못 얹은 입금은 반환."""
    rest=[]
    for d in sorted(leftover, key=lambda x:x['date']):
        tg=_net_unpaid_rows(ws, C)
        placed=False
        for r,_,_ in tg:
            if ws.cell(r,C['입금']).value in (None,"") and ws.cell(r,C['은행']).value in (None,""):
                ws.cell(r,C['은행'], d.get('bank',''))
                ws.cell(r,C['입금일'], (f"{d['date'].isoformat()}\n({d['maturity']})" if d.get('maturity') else d['date']))
                ws.cell(r,C['입금'], d['amt'])
                jc=ws.cell(r,C['차액']).value
                if not (isinstance(jc,str) and jc.startswith("=")):
                    ws.cell(r,C['차액'], f"={EL(C['정산'])}{r}-{EL(C['입금'])}{r}")
                placed=True; break
        if not placed: rest.append(d)
    return rest

def unpaid_after_credit(ws, C, exc=None):
    """계산서별 미수 계산 후, 과입금(음수=초과수령)+미배분 floating 입금을 부족분(미수)에 오래된 순으로 상계.
    반환: 상계 후에도 1000원 초과로 남는 [(작성일, 미수액)]."""
    ups=compute_unpaid(ws, C, exc)
    credit=sum(-u for d,u in ups if u<0) + loose_credit(ws, C)   # 과입금 + 미배분 입금 = 신용 잔액
    short=sorted([[d,u] for d,u in ups if u>1000], key=lambda x:x[0])
    for it in short:
        use=min(it[1],credit); it[1]-=use; credit-=use
    return [(d,u) for d,u in short if u>1000]

def clean_filename_name(name):
    n=LEGAL.sub("", str(name))
    n=re.sub(r"\s+"," ",n).strip()
    n=re.sub(r"\(\s*\)","",n).strip()   # 빈 괄호만 제거(앞의 (자) 같은 정상 괄호는 보존)
    return n

# ---------- 누적 ----------
def read_existing(ws):
    C=cols(ws); invs=[]; deps=[]
    for r in range(3,ws.max_row+1):
        ad=pdate(ws.cell(r,C['작성']).value); amt=ws.cell(r,C['합계']).value
        if not isinstance(amt,(int,float)):
            sup=ws.cell(r,C['공급']).value
            amt=round(sup*1.1) if isinstance(sup,(int,float)) else None
        if ad and isinstance(amt,(int,float)):
            invs.append({'row':r,'date':ad,'amt':int(amt),'paid':0})
        dd=pdate(ws.cell(r,C['입금일']).value); dm=ws.cell(r,C['입금']).value
        if dd and isinstance(dm,(int,float)):
            deps.append({'row':r,'date':dd,'amt':int(dm)})
    return invs,deps

def rebuild_clean(ws):
    """데이터 범위만 새 워크북에 복사 + 폰트통일 + 빈행제거"""
    C=cols(ws); maxc=C['n']
    # 마지막 데이터행
    L=last_data_row(ws,C)
    nb=Workbook(); nws=nb.active
    src_rows=[r for r in range(1,L+1)]
    merges=[str(m) for m in ws.merged_cells.ranges]
    out_r=0
    rowmap={}
    for r in src_rows:
        # 데이터행(3행 이상)에서 완전 빈 행은 건너뜀
        if r>2 and all(ws.cell(r,c).value in (None,"") for c in range(1,maxc+1)): 
            continue
        out_r+=1; rowmap[r]=out_r
        for c in range(1,maxc+1):
            s=ws.cell(r,c); d=nws.cell(out_r,c)
            if isinstance(s,MergedCell): continue
            d.value=s.value
            d.number_format=s.number_format
            d.alignment=copy(s.alignment)
    for mg in merges:
        m=re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)",mg)
        if m and int(m.group(2)) in rowmap and int(m.group(4)) in rowmap:
            try: nws.merge_cells(f"{m.group(1)}{rowmap[int(m.group(2))]}:{m.group(3)}{rowmap[int(m.group(4))]}")
            except: pass
    apply_font(nws); autofit(nws,maxc)
    return nb

def new_template():
    wb=Workbook(); ws=wb.active
    # 가능유공압 표준 헤더: 차액·비고 라벨은 1행, J1:J2·K1:K2 병합
    ws.cell(1,1,"전자세금계산서"); ws.cell(1,7,"은행입금내역"); ws.cell(1,10,"차액"); ws.cell(1,11,"비고")
    for i,h in enumerate(["작성일","발행처","공급가","부가세","합계","정산액","은행","입금일","입금액"],1):
        ws.cell(2,i,h)
    try:
        ws.merge_cells("A1:F1"); ws.merge_cells("G1:I1"); ws.merge_cells("J1:J2"); ws.merge_cells("K1:K2")
    except: pass
    return wb

def _rebuild_loose_tail(ws, C, leftover, issuer_col, EL):
    """바닥의 자기참조-전용 미정산 구간을 (미배분입금 포함) 날짜순으로 재배치."""
    merged=set()
    for mg in ws.merged_cells.ranges:
        mm=re.match(r"[A-Z]+(\d+):[A-Z]+(\d+)",str(mg))
        if mm: merged.update(range(int(mm.group(1)),int(mm.group(2))+1))
    def loose_inv(r):
        if r in merged: return False
        if not pdate(ws.cell(r,C['작성']).value): return False
        if ws.cell(r,C['입금']).value not in (None,""): return False
        if ws.cell(r,C['은행']).value not in (None,""): return False
        if ws.cell(r,C['비고']).value not in (None,""): return False
        jc=ws.cell(r,C['정산']).value
        if jc not in (None,"") and jc!=f"={EL(C['합계'])}{r}": return False
        return True
    def dep_only(r):
        if r in merged: return False
        if pdate(ws.cell(r,C['작성']).value): return False
        return ws.cell(r,C['입금']).value not in (None,"")
    start=None; r=ws.max_row
    while r>=3 and (loose_inv(r) or dep_only(r)):
        start=r; r-=1
    if start is None and not leftover: return
    if start is None: start=ws.max_row+1
    items=[]
    for rr in range(start, ws.max_row+1):
        if loose_inv(rr):
            items.append(('inv', pdate(ws.cell(rr,C['작성']).value), ws.cell(rr,C['공급']).value,
                          ws.cell(rr,C['발행']).value if issuer_col else None, None,None,None))
        elif dep_only(rr):
            items.append(('dep', pdate(ws.cell(rr,C['입금일']).value), None,None,
                          ws.cell(rr,C['은행']).value, ws.cell(rr,C['입금일']).value, ws.cell(rr,C['입금']).value))
    for d in leftover:
        dl=f"{d['date'].isoformat()}\n({d['maturity']})" if d.get('maturity') else d['date']
        items.append(('dep', d['date'], None,None, d.get('bank',''), dl, d['amt']))
    items.sort(key=lambda x:(x[1], 0 if x[0]=='inv' else 1))
    end=ws.max_row+len(leftover)
    for rr in range(start, end+1):
        for c in range(1,C['n']+1):
            cc=ws.cell(rr,c)
            if not isinstance(cc,MergedCell): cc.value=None
    rr=start
    for it in items:
        if it[0]=='inv':
            ws.cell(rr,C['작성'],it[1])
            if issuer_col and it[3] is not None: ws.cell(rr,C['발행'],it[3])
            ws.cell(rr,C['공급'],it[2])
            ws.cell(rr,C['부가'],f"={EL(C['공급'])}{rr}*0.1")
            ws.cell(rr,C['합계'],f"={EL(C['공급'])}{rr}+{EL(C['부가'])}{rr}")
            ws.cell(rr,C['정산'],f"={EL(C['합계'])}{rr}")
            ws.cell(rr,C['차액'],f"={EL(C['정산'])}{rr}-{EL(C['입금'])}{rr}")
        else:
            ws.cell(rr,C['은행'],it[4]); ws.cell(rr,C['입금일'],it[5]); ws.cell(rr,C['입금'],it[6])
            ws.cell(rr,C['차액'], -int(it[6]) if isinstance(it[6],(int,float)) else None)   # 미배분 입금 = 과입금(-입금), 고정값
        rr+=1

def accumulate(existing_path, new_invoices, new_deposits, issuer, out_path, exceptions=None):
    """원본 정산은 그대로 보존(제자리), 신규 계산서 추가 + 신규 입금만 미정산 계산서에 채움."""
    EL=get_column_letter
    flags=[]
    if existing_path and os.path.exists(existing_path):
        wb=load_workbook(existing_path); ws=wb.active
        wv=load_workbook(existing_path,data_only=True).active
        C=cols(ws)
        for row in ws.iter_rows(min_row=1,max_row=2):
            for c in row:
                if not isinstance(c,MergedCell) and str(c.value).strip()=="월계": c.value="정산액"
        # 비고 헤더 보장(이미 있으면/병합셀이면 건드리지 않음)
        bgc=ws.cell(2,C['비고'])
        has_bigo=(str(ws.cell(1,C['비고']).value).strip()=="비고") or (str(bgc.value).strip()=="비고")
        if not has_bigo and not isinstance(bgc,MergedCell): bgc.value="비고"
        # 트레일링 빈행 제거
        L=last_data_row(ws,C)
        if ws.max_row>L: ws.delete_rows(L+1, ws.max_row-L)
        # 정산완료(입금 있는 묶음) 행
        mranges=[]
        for mg in ws.merged_cells.ranges:
            mm=re.match(r"[A-Z]+(\d+):[A-Z]+(\d+)",str(mg))
            if mm: mranges.append((int(mm.group(1)),int(mm.group(2))))
        settled=set()
        for r in range(3,ws.max_row+1):
            dm=wv.cell(r,C['입금']).value
            if isinstance(dm,(int,float)) and dm!=0 and pdate(wv.cell(r,C['입금일']).value):
                settled.add(r)
                for (a,b) in mranges:
                    if a<=r<=b: settled.update(range(a,b+1))
        # 미정산 계산서
        unsettled=[]
        for r in range(3,ws.max_row+1):
            if isinstance(ws.cell(r,C['공급']),MergedCell): continue
            a=pdate(ws.cell(r,C['작성']).value); sup=wv.cell(r,C['공급']).value
            if a and isinstance(sup,(int,float)) and r not in settled:
                unsettled.append({'row':r,'date':a,'amt':int(round(sup*1.1))})
        # 기존 계산서 (날짜,공급가) 중복방지용
        from collections import Counter as _Counter
        exist_cnt=_Counter()
        for r in range(3,ws.max_row+1):
            if isinstance(ws.cell(r,C['공급']),MergedCell): continue
            a=pdate(ws.cell(r,C['작성']).value); sup=wv.cell(r,C['공급']).value
            if a and isinstance(sup,(int,float)): exist_cnt[(a,int(round(sup)))]+=1
        issuer_col=C['발행'] is not None
    else:
        wb=new_template(); ws=wb.active; C=cols(ws)
        unsettled=[]; issuer_col=True
        from collections import Counter as _Counter
        exist_cnt=_Counter()
    # 신규 계산서 추가 (마지막 데이터행 다음, 중복 제외)
    cur=last_data_row(ws,C); _added_inv=0
    for inv in sorted(new_invoices,key=lambda x:x['date']):
        k=(inv['date'],int(round(inv['sup'])))
        if exist_cnt.get(k,0)>0: exist_cnt[k]-=1; continue
        cur+=1; r=cur; _added_inv+=1
        ws.cell(r,C['작성'],inv['date'])
        if issuer_col: ws.cell(r,C['발행'],issuer)
        ws.cell(r,C['공급'],inv['sup'])
        ws.cell(r,C['부가'],f"={EL(C['공급'])}{r}*0.1")
        ws.cell(r,C['합계'],f"={EL(C['공급'])}{r}+{EL(C['부가'])}{r}")
        ws.cell(r,C['정산'],f"={EL(C['합계'])}{r}")
        ws.cell(r,C['차액'],f"={EL(C['정산'])}{r}-{EL(C['입금'])}{r}")
        unsettled.append({'row':r,'date':inv['date'],'amt':inv['amt']})
    # === 신규 입금 매칭(정확: 단일/조합) → 미정산 계산서 ===
    import itertools
    unsettled.sort(key=lambda x:(x['date'],x['row']))
    uamt=[u['amt'] for u in unsettled]
    unm=set(range(len(unsettled)))
    deps=sorted([dict(d) for d in new_deposits],key=lambda d:d['date'])
    NEAR=1000   # 근사매칭 허용오차: 차액 1000원 이하는 미수로 안 잡는 규칙과 연동
    def _find(target,size,ddate,tol):
        cands=sorted([i for i in unm if unsettled[i]['date']<=ddate+timedelta(days=45)],key=lambda i:unsettled[i]['date'])
        if len(cands)>22: cands=cands[-22:]
        best=None; bd=tol+1
        for combo in itertools.combinations(cands,size):
            diff=abs(sum(uamt[i] for i in combo)-target)
            if diff<=tol and diff<bd: best=list(combo); bd=diff
            if bd==0: break
        return best
    def _run(d,tol):  # 연속 구간 합(n:1 큰 묶음, ≥4개)
        idxs=[i for i in range(len(unsettled)) if i in unm and unsettled[i]['date']<=d["date"]+timedelta(days=45)][-50:]
        best=None; bd=tol+1
        for L in range(len(idxs)):
            s=0
            for R in range(L,len(idxs)):
                s+=uamt[idxs[R]]
                if (R-L+1)>=4 and abs(s-d['amt'])<=tol and abs(s-d['amt'])<bd:
                    best=idxs[L:R+1]; bd=abs(s-d['amt'])
        return best
    settlements=[]; matched=set()
    for tol in (2, NEAR):     # 정확매칭(±2) 전부 먼저 → 그 다음 근사매칭(±1000)
        for size in (1,2,3):
            for dj,d in enumerate(deps):
                if dj in matched: continue
                sub=_find(d['amt'],size,d['date'],tol)
                if sub:
                    settlements.append({'inv':sub,'dep':d}); matched.add(dj)
                    for i in sub: unm.discard(i)
        for dj,d in enumerate(deps):
            if dj in matched: continue
            found=_run(d,tol)
            if found:
                settlements.append({'inv':found,'dep':d}); matched.add(dj)
                for i in found: unm.discard(i)
    leftover=[deps[dj] for dj in range(len(deps)) if dj not in matched]
    def put_dep(rr,d):
        ws.cell(rr,C['은행'],d.get('bank','')); ws.cell(rr,C['입금'],d['amt'])
        if d.get('maturity'): ws.cell(rr,C['입금일'],f"{d['date'].isoformat()}\n({d['maturity']})")
        else: ws.cell(rr,C['입금일'],d['date'])
    # === 채우기 ===
    for st in settlements:
        sub=sorted(st['inv'],key=lambda i:unsettled[i]['row']); d=st['dep']
        rows=[unsettled[i]['row'] for i in sub]; anchor=rows[0]
        put_dep(anchor,d)
        if len(rows)==1: ws.cell(anchor,C['정산'],f"={EL(C['합계'])}{anchor}")
        elif rows==list(range(rows[0],rows[-1]+1)): ws.cell(anchor,C['정산'],f"=SUM({EL(C['합계'])}{rows[0]}:{EL(C['합계'])}{rows[-1]})")
        else: ws.cell(anchor,C['정산'],"="+"+".join(f"{EL(C['합계'])}{rr}" for rr in rows))
        ws.cell(anchor,C['차액'],f"={EL(C['정산'])}{anchor}-{EL(C['입금'])}{anchor}")
        contiguous=(rows==list(range(rows[0],rows[-1]+1)) and len(rows)>1)
        if contiguous:
            for col in [C['정산'],C['차액'],C['은행'],C['입금일'],C['입금'],C['비고']]:
                try: ws.merge_cells(start_row=anchor,start_column=col,end_row=rows[-1],end_column=col)
                except: pass
        else:
            for rr in rows[1:]:
                ws.cell(rr,C['은행'],"합산정산"); ws.cell(rr,C['입금일'],d['date'])
                ws.cell(rr,C['입금']).value=None; ws.cell(rr,C['정산']).value=None
                ws.cell(rr,C['차액'],0); ws.cell(rr,C['비고'],f"{anchor}행")
    # 미배분 입금 → 바닥 loose 구간에 날짜순 끼워넣기(자기참조 행만 안전 재배치)
    _rebuild_loose_tail(ws, C, leftover, issuer_col, EL)
    for d in leftover: flags.append(f"미배분입금 {d['date']} {d['amt']:,}")
    # 예외처리(차액 0)
    if exceptions:
        for r in range(3,ws.max_row+1):
            a=pdate(ws.cell(r,C['작성']).value); sup=ws.cell(r,C['공급']).value
            if a and isinstance(sup,(int,float)):
                if (a.isoformat(),round(sup*1.1)) in exceptions or (a.isoformat(),int(sup)) in exceptions:
                    jc=ws.cell(r,C['차액'])
                    if not isinstance(jc,MergedCell): jc.value=0
    apply_jeolsang(ws,C)
    apply_formats(ws,C); apply_font(ws); autofit(ws,C['n']); ws.freeze_panes='A3'
    wb.save(out_path)
    return dict(new_inv=_added_inv, new_dep=len(new_deposits), flags=flags)

def validate(path):
    """구조 점검: '월계' 헤더 잔존, 데이터 중간 공백행."""
    iss=[]; ws=load_workbook(path).active; C=cols(ws)
    for r in (1,2):
        for c in range(1,ws.max_column+1):
            if str(ws.cell(r,c).value).strip()=="월계": iss.append("'월계' 헤더 잔존")
    L=last_data_row(ws,C)
    for r in range(3,L+1):
        if all(ws.cell(r,c).value in (None,"") for c in range(1,C['n']+1)):
            iss.append(f"{r}행 공백")
    return iss

# ===== 엔진 끝 (truncation 방지 여백) =====
# buffer 1
# buffer 2
# buffer 3
