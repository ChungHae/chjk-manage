# -*- coding: utf-8 -*-
"""웹앱 처리 파이프라인: 업로드 파일 자동판별(매출/매입/어음/은행) → 기존 누적본에 신규만 추가 → 양식대로 출력.
run.py 로직을 동적 입력용으로 재구성. 지역(서울/화성)은 업로드 칸으로 구분해 받음."""
import os, re, glob, shutil, subprocess, tempfile, types
from datetime import date
from collections import defaultdict, Counter
import pandas as pd
from openpyxl import load_workbook, Workbook

_HERE = os.path.dirname(os.path.abspath(__file__))
_ep = os.path.join(_HERE, "engine_core.py")
if not os.path.exists(_ep): _ep = os.path.join(_HERE, "engine", "engine.py")
E = types.ModuleType("engine"); exec(compile(open(_ep, encoding="utf-8").read(), _ep, "exec"), E.__dict__)

# ---------- 파일 변환·판별 ----------
def to_xlsx(path, workdir):
    """모든 업로드를 깨끗한 .xlsx로 변환(.xls·스타일깨진 .xlsx 대응). LibreOffice 사용, 실패 시 원본."""
    base = os.path.splitext(os.path.basename(path))[0]
    out = os.path.join(workdir, base + ".xlsx")
    try:
        subprocess.run(["libreoffice", "--headless", "--convert-to", "xlsx", "--outdir", workdir, path],
                       check=True, capture_output=True, timeout=90)
        if os.path.exists(out): return out
    except Exception:
        pass
    # 변환 실패 시: 이미 xlsx면 그대로
    if path.lower().endswith(".xlsx"):
        shutil.copy(path, out); return out
    return path

def detect_type(path):
    """헤더로 파일 종류 판별: '어음'/'은행'/'매출'/'매입'/None."""
    try:
        raw = pd.read_excel(path, header=None, dtype=object, nrows=20)
    except Exception:
        return None
    flat = " ".join(str(x) for x in raw.values.flatten() if pd.notna(x))
    if ("전자어음번호" in flat) or ("만기일자" in flat and "발행인" in flat) \
       or ("판매채권" in flat) or ("채권번호" in flat) \
       or ("만기일" in flat and ("구매기업" in flat or "채권금액" in flat)):
        return "어음"
    if ("공급자사업자등록번호" in flat) and ("공급받는자사업자등록번호" in flat):
        # 매출=충해전기가 공급자(값 고정) / 매입=충해전기가 공급받는자(값 고정). 한쪽이 거의 1개 값이면 그쪽이 우리.
        try:
            df = pd.read_excel(path, header=None, dtype=object, nrows=400)
            hr = None
            for i in range(min(15, len(df))):
                row = " ".join(str(x) for x in df.iloc[i].tolist())
                if "공급자사업자등록번호" in row and "공급받는자사업자등록번호" in row: hr = i; break
            if hr is not None:
                cols = [str(x) for x in df.iloc[hr].tolist()]
                sup_i = next((j for j, c in enumerate(cols) if "공급자사업자등록번호" in c and "받는" not in c), None)
                buy_i = next((j for j, c in enumerate(cols) if "공급받는자사업자등록번호" in c), None)
                body = df.iloc[hr + 1:]
                def _uniq(ci):
                    if ci is None: return 9999
                    vals = [str(v).strip() for v in body.iloc[:, ci].tolist() if str(v).strip() not in ("", "nan", "None")]
                    return len(set(vals)) if vals else 9999
                su, bu = _uniq(sup_i), _uniq(buy_i)
                return "매출" if su <= bu else "매입"
        except Exception:
            pass
        return "매출"
    if ("거래일" in flat or "거래일시" in flat) and ("입금" in flat):
        return "은행"
    return None

def _col_max_date(path, keys, today=None):
    """헤더에서 keys(거래일/작성일 등) 컬럼을 찾아, 그 컬럼 데이터의 가장 늦은 날짜(오늘 이후 제외)."""
    import datetime as _dt
    today = today or _dt.date.today()
    try: raw = pd.read_excel(path, header=None, dtype=object, nrows=100000)
    except Exception: return None
    col = None; hr = None
    for i in range(min(20, len(raw))):
        for j, v in enumerate(raw.iloc[i].tolist()):
            sv = str(v).replace(" ", "")
            if any(k in sv for k in keys): hr = i; col = j; break
        if col is not None: break
    if col is None: return None
    best = None
    for i in range(hr + 1, len(raw)):
        m = re.search(r"(20\d{2})[.\-/](\d{1,2})[.\-/](\d{1,2})", str(raw.iloc[i, col]))
        if not m: continue
        try: d = _dt.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError: continue
        if d > today: continue
        s = d.isoformat()
        if best is None or s > best: best = s
    return best

# ---------- 누적 처리 ----------
def process(uploads_by_loc, data_dir, out_dir, progress=None):
    """uploads_by_loc: {'서울':[paths...], '화성':[paths...]} (지역별 업로드 파일 경로).
    data_dir: 기존 누적본(출력 xlsx들 + 지원표) 폴더. out_dir: 결과 출력 폴더.
    반환: dict(summary=[...], detected={...}, status=Counter, report_path, out_dir)."""
    def log(m):
        if progress: progress(m)
    work = tempfile.mkdtemp()
    # 1) 업로드 변환 + 판별
    detected = {loc: {"매출": [], "매입": [], "어음": [], "은행": [], "미상": []} for loc in uploads_by_loc}
    for loc, paths in uploads_by_loc.items():
        for p in paths:
            x = to_xlsx(p, work); t = detect_type(x) or "미상"
            detected[loc][t].append(x)
            log(f"{loc} 판별: {os.path.basename(p)} → {t}")
    # 2) 지원표 로드
    def _load_tbl(name, sheet, cols):
        p = os.path.join(data_dir, name); rows = []
        if os.path.exists(p):
            wb = load_workbook(p); ws = wb[sheet] if sheet in wb.sheetnames else wb.active
            for r in range(2, ws.max_row+1): rows.append([ws.cell(r, c).value for c in cols])
        return rows
    alias = {}
    for loc, nm, tg in [(r[0], r[1], r[2]) for r in _load_tbl("_거래처별칭표.xlsx", "Sheet", [1, 2, 3])]:
        if tg and str(tg).strip().lower() != "none": alias[(loc, E.cname(nm))] = str(tg).split(",")[0].strip()
    duerules = {}
    for r in _load_tbl("_거래처기준설정표.xlsx", "기준설정", [3, 4]):
        if r[0] and r[1]: duerules[str(r[0]).strip()] = str(r[1]).strip()
    exceptions = defaultdict(set)
    for r in _load_tbl("_예외처리표.xlsx", "예외처리", [3, 4, 5]):
        if r[0] and r[1] and r[2]:
            ad = E.pdate(r[1])
            if ad: exceptions[str(r[0]).strip()].add((ad.isoformat(), int(r[2])))
    excluded = set(str(r[0]).strip() for r in _load_tbl("_제외거래처표.xlsx", "제외거래처", [3]) if r[0])
    reassign = {}
    for r in _load_tbl("_입금재배정표.xlsx", "입금재배정", [1, 2, 4, 8]):
        ad = E.pdate(r[0])
        if ad and r[1] and r[2] and r[3]:
            reassign[(ad.isoformat(), int(r[1]))] = (str(r[3]).strip(), str(r[2]).strip())
    # 3) 기존 누적본 인덱스
    exist = {}; universe = defaultdict(dict)
    for f in glob.glob(os.path.join(data_dir, "서울", "*.xlsx")) + glob.glob(os.path.join(data_dir, "화성", "*.xlsx")):
        b = os.path.basename(f)
        if b.startswith("_") or "~$" in b: continue
        m = re.search(r"\((\d{3}-\d{2}-\d{5})\)", b); 
        if not m: continue
        loc = "서울" if (os.sep+"서울"+os.sep in f) else "화성"
        bn = m.group(1)
        if (loc, bn) in exist: continue
        nm = re.sub(r"^[^)]*\)\s*", "", b).rsplit(" (", 1)[0]
        exist[(loc, bn)] = f; universe[loc][E.cname(nm)] = nm
    # 4) 매출/매입 파싱
    sales = {loc: [] for loc in uploads_by_loc}; sales_tot = defaultdict(int); purch_tot = defaultdict(int); names = {}
    for loc in uploads_by_loc:
        for fp in detected[loc]["매출"]:
            for x in E.parse_sales(fp):
                sales[loc].append(x); sales_tot[x["사업자번호"]] += x["합계"]; names[x["사업자번호"]] = x["거래처"]
                universe[loc].setdefault(E.cname(x["거래처"]), x["거래처"])
        for fp in detected[loc]["매입"]:
            for x in E.parse_purchase(fp): purch_tot[x["사업자번호"]] += x["합계"]
    skip = {bn for bn in sales_tot if purch_tot.get(bn, 0) > sales_tot[bn] and purch_tot.get(bn, 0) > 0} | excluded
    # 5) nb_map
    nb_map = defaultdict(dict)
    for (loc, bn), f in exist.items():
        nm = re.sub(r"^[^)]*\)\s*", "", os.path.basename(f)).rsplit(" (", 1)[0]; nb_map[loc][E.cname(nm)] = bn
    for loc in uploads_by_loc:
        for x in sales[loc]: nb_map[loc].setdefault(E.cname(x["거래처"]), x["사업자번호"])
    def _bn_of(loc, who):
        if not who: return None
        wc = E.cname(who); m = nb_map[loc].get(wc)
        if m: return m
        cand = set(v for k, v in nb_map[loc].items() if k and len(k) >= 2 and (k.startswith(wc) or wc.startswith(k)))
        return next(iter(cand)) if len(cand) == 1 else None
    # 6) 입금(은행) + 어음 → 거래처 배정
    cust_dep = defaultdict(list); unassigned = []
    for loc in uploads_by_loc:
        for fp in detected[loc]["은행"]:
            _bnk = next((b for b in ["신한", "기업", "하나", "농협", "우리", "국민", "외환", "수협", "산업", "씨티", "케이뱅크", "카카오", "토스"] if b in os.path.basename(fp)), "은행")
            for d in E.parse_bank(fp, _bnk):
                _ra = reassign.get((d["거래일"], int(d["입금액"])))
                if _ra: cust_dep[_ra].append(d); continue
                kind, who = E.match_deposit(loc, d["상대방명"], universe, alias)
                if kind == "제외": continue
                bn = _bn_of(loc, who)
                if not bn: unassigned.append((loc, d["거래일"], d["입금액"], d["상대방명"], kind or "미배정")); continue
                cust_dep[(loc, bn)].append(d)
        for fp in detected[loc]["어음"]:
            _bnk = next((b for b in ["하나", "신한", "기업", "농협", "우리", "국민"] if b in os.path.basename(fp)), "신한")
            for e in E.parse_eum(fp, bank=_bnk):
                _d = {"거래일": e["수취일"], "입금액": e["어음금액"], "상대방명": e["발행인"], "은행": e.get("bank", ""), "만기": e["만기일"]}
                _ra = reassign.get((e["수취일"], int(e["어음금액"])))
                if _ra: cust_dep[_ra].append(_d); continue
                kind, who = E.match_deposit(loc, e["발행인"], universe, alias)
                bn = _bn_of(loc, who)
                if not bn: unassigned.append((loc, e["수취일"], e["어음금액"], e["발행인"]+"(어음)", "어음-미배정")); continue
                cust_dep[(loc, bn)].append({"거래일": e["수취일"], "입금액": e["어음금액"], "상대방명": e["발행인"], "은행": e.get("bank", ""), "만기": e["만기일"]})
    # 7) 거래처별 누적 출력
    cust_inv = defaultdict(list)
    for loc in uploads_by_loc:
        for x in sales[loc]:
            cust_inv[(loc, x["사업자번호"])].append({"date": date(*map(int, x["작성일"].split("-"))),
                "sup": x["공급가"] if x["공급가"] is not None else round(x["합계"]/1.1), "amt": x["합계"]})
    summary = []; flags = []
    PRE = {"장기미수": "🔴장기미수", "미수": "🟠미수", "진행": "🟡진행", "완납": "✅완납"}
    for loc in uploads_by_loc:
        os.makedirs(os.path.join(out_dir, loc), exist_ok=True)
        bset = set(bn for (l, bn) in cust_inv if l == loc) | set(bn for (l, bn) in cust_dep if l == loc) | set(bn for (l, bn) in exist if l == loc)
        for bn in bset:
            if bn in skip: continue
            path = exist.get((loc, bn)); name = names.get(bn) or (re.sub(r"^[^)]*\)\s*", "", os.path.basename(path)).rsplit(" (", 1)[0] if path else bn)
            inv = cust_inv.get((loc, bn), [])
            ek = set()
            if path:
                wv = load_workbook(path).active; Cv = E.cols(wv)
                for r in range(3, wv.max_row+1):
                    dd = E.pdate(wv.cell(r, Cv["입금일"]).value); dm = wv.cell(r, Cv["입금"]).value
                    if dd and isinstance(dm, (int, float)): ek.add((dd.isoformat(), int(dm)))
            nd = []
            for d in cust_dep.get((loc, bn), []):
                di = d["거래일"]
                if (di, int(d["입금액"])) in ek: continue
                ent = {"date": date(*map(int, di.split("-"))), "amt": int(d["입금액"]), "bank": d.get("은행", "")}
                if d.get("만기"): ent["maturity"] = d["만기"]
                nd.append(ent)
            if not inv and not nd and not path: continue
            tmp = os.path.join(work, "_o.xlsx"); exc = exceptions.get(bn, set())
            try: res = E.accumulate(path, inv, nd, loc, tmp, exc)
            except Exception as ex: flags.append((loc, name, f"오류:{ex}")); continue
            wsf = load_workbook(tmp).active; Cf = E.cols(wsf); worst = "완납"; out_amt = 0; earliest = None
            supmap = {}
            for r in range(3, wsf.max_row + 1):
                da = E.pdate(wsf.cell(r, Cf['작성']).value); sv = wsf.cell(r, Cf['공급']).value
                if da and isinstance(sv, (int, float)): supmap[da.isoformat()] = supmap.get(da.isoformat(), 0) + sv
            for a, bal in E.unpaid_after_credit(wsf, Cf, exc):
                st = E.status(a, rule=E.resolve_rule(duerules.get(bn, "익월말"), supmap.get(a.isoformat(), 0)))
                if st in ("미수", "장기미수"):
                    out_amt += bal
                    if earliest is None or a < earliest: earliest = a
                worst = "장기미수" if st == "장기미수" else ("미수" if st == "미수" and worst != "장기미수" else ("진행" if st == "기한내" and worst == "완납" else worst))
            cn = E.clean_filename_name(name)
            fn = re.sub(r'[\\/:*?"<>|]', "", f"{PRE[worst]} {loc}) {cn} ({bn}).xlsx")
            dest = os.path.join(out_dir, loc, fn)
            shutil.move(tmp, dest)
            summary.append((loc, name, bn, worst, out_amt, earliest.isoformat() if earliest else "", res["new_inv"], res["new_dep"], "신규" if not path else "기존"))
    # 미처리 기존본은 그대로 복사(누적 유지)
    for (loc, bn), f in exist.items():
        if bn in skip: continue
        if not any(s[2] == bn and s[0] == loc for s in summary):
            d = os.path.join(out_dir, loc, os.path.basename(f)); os.makedirs(os.path.dirname(d), exist_ok=True)
            if not os.path.exists(d): shutil.copy(f, d)
    # 지원표도 출력 폴더로 복사
    for tbl in glob.glob(os.path.join(data_dir, "_*.xlsx")): shutil.copy(tbl, os.path.join(out_dir, os.path.basename(tbl)))
    status = Counter(s[3] for s in summary)
    new_companies = [s[1] for s in summary if s[8] == "신규"]
    # 기준일 = 업로드 자료의 최근 거래일/작성일(은행·세금계산서 스캔 + 어음 수취일; 미래 만기 제외)
    _bc = []
    for loc in uploads_by_loc:
        for fp in detected[loc]["은행"]:
            d = _col_max_date(fp, ["거래일시", "거래일", "거래일자"])
            if d: _bc.append(d)
        for fp in detected[loc]["매출"] + detected[loc]["매입"]:
            d = _col_max_date(fp, ["작성일자", "작성일"])
            if d: _bc.append(d)
        for fp in detected[loc]["어음"]:
            try:
                for e in E.parse_eum(fp):
                    if e.get("수취일"): _bc.append(e["수취일"])
            except Exception: pass
    basis = max(_bc) if _bc else None
    return dict(summary=summary, detected=detected, status=dict(status), unassigned=unassigned,
                flags=flags, out_dir=out_dir, new_companies=new_companies, basis=basis)


# ---------- 갱신본 검증 (손상 + 입금누락) ----------
def _bizno(fn):
    m = re.search(r'\((\d{3}-\d{2}-\d{5})\)', fn); return m.group(1) if m else fn

def _deps_count(f):
    ws = load_workbook(f).active; C = E.cols(ws); n = 0
    for r in range(3, ws.max_row + 1):
        dd = E.pdate(ws.cell(r, C['입금일']).value); dm = ws.cell(r, C['입금']).value
        if dd and isinstance(dm, (int, float)) and dm > 0: n += 1
    return n

def verify(out_dir, old_dir):
    """갱신본 검증: 파일 손상 + 직전 대비 입금 감소(누락). returns (ok, problems[])"""
    probs = []; newd = {}
    for f in glob.glob(out_dir + '/서울/*.xlsx') + glob.glob(out_dir + '/화성/*.xlsx'):
        bn = os.path.basename(f)
        if bn.startswith(('_', '~$')): continue
        loc = '서울' if (os.sep + '서울' + os.sep) in f else '화성'
        try: newd[(loc, _bizno(bn))] = _deps_count(f)
        except Exception: probs.append('손상: ' + bn)
    for f in glob.glob(old_dir + '/서울/*.xlsx') + glob.glob(old_dir + '/화성/*.xlsx'):
        bn = os.path.basename(f)
        if bn.startswith(('_', '~$')): continue
        loc = '서울' if (os.sep + '서울' + os.sep) in f else '화성'
        try: o = _deps_count(f)
        except Exception: continue
        n = newd.get((loc, _bizno(bn)))
        if n is not None and n < o: probs.append(f'입금감소: {bn} ({o}->{n})')
    return (len(probs) == 0, probs)
