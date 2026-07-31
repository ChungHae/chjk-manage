# -*- coding: utf-8 -*-
import os, io, zipfile, tempfile, glob, datetime, re
import hashlib, requests, json
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import pipeline, store

st.set_page_config(page_title="충해전기 관리시스템", page_icon="📒", layout="wide")
import streamlit.components.v1 as _components
_components.html("<script>window.parent.document.title='충해전기 관리시스템';</script>", height=0)
st.markdown("<style>[data-testid='stMain']{scrollbar-gutter:stable;}.block-container{max-width:1320px;margin:0 auto;padding-top:1.4rem;padding-left:3rem;padding-right:3rem;}@media(max-width:640px){.block-container{padding-left:0.7rem!important;padding-right:0.7rem!important;}}</style><style>iframe[height='0']{display:none;}.st-key-nytrig{position:fixed!important;left:-9999px!important;top:0!important;width:1px!important;height:1px!important;overflow:hidden!important;}a[href$='page_nyung'],[data-testid='stTopNavLink'][href$='page_nyung'],[data-testid='stTopNavLinkContainer']:has(a[href$='page_nyung']),li:has(a[href$='page_nyung']),[data-testid='stSidebarNavItems'] li:has(a[href$='page_nyung']){display:none!important;}</style>", unsafe_allow_html=True)
st.markdown("""<style>
/* 작은 화면(≤768px)에서 Streamlit은 상단 탭 대신 네비게이션을 사이드바(햄버거)로 옮긴다.
   그 사이드바 네비를 다시 상단 가로바로 꺼내 모든 카테고리를 PC처럼 상단에 표시한다. */
@media (max-width: 768px){
  /* 사이드바(드로어)를 폭 0으로 강제 고정 → 네비 클릭 시 드로어가 열려도 화면을 덮지 않음 */
  section[data-testid="stSidebar"]{ transform:none !important; width:0 !important; min-width:0 !important; max-width:0 !important; }
  [data-testid="stSidebarContent"]{ overflow:visible !important; }
  [data-testid="stSidebarNav"]{
    position:fixed !important; top:0 !important; left:0 !important; right:0 !important;
    width:100vw !important; max-width:100vw !important; box-sizing:border-box !important;
    z-index:2147483000 !important; background:#fff !important;
    border-bottom:1px solid #d6e0ec !important; box-shadow:0 1px 5px rgba(0,0,0,.12) !important;
    padding:5px 6px !important; margin:0 !important;
  }
  [data-testid="stSidebarNavItems"]{
    display:flex !important; flex-direction:row !important; flex-wrap:wrap !important;
    justify-content:center !important; gap:4px !important;
    max-height:none !important; padding:0 !important; margin:0 !important; list-style:none !important;
  }
  [data-testid="stSidebarNavItems"] li, [data-testid="stSidebarNavLinkContainer"]{ width:auto !important; margin:0 !important; }
  [data-testid="stSidebarNavLink"]{
    width:auto !important; padding:5px 10px !important; white-space:nowrap !important;
    border:1px solid #e0e6ee !important; border-radius:7px !important;
  }
  [data-testid="stExpandSidebarButton"]{ display:none !important; }
  /* 상단의 숨은 주입용 요소(height=0 iframe·<style> 마크다운)가 세로 gap을 먹어 로고까지 공백이 큼 → 접기 */
  [data-testid="stMainBlockContainer"] [data-testid="stElementContainer"]:has(iframe[scrolling="no"]),
  [data-testid="stMainBlockContainer"] [data-testid="stElementContainer"]:has(> [data-testid="stMarkdown"] style){ display:none !important; }
  [data-testid="stMainBlockContainer"]{ padding-top:58px !important; }
  /* 네이티브 header()(자료처리 등) 로고·제목을 미수현황 모바일 크기(로고 40·제목 18·부제 11)에 맞춤 */
  .pg-hdr{ gap:10px !important; margin:6px 0 14px !important; }
  .pg-hdr-logo{ height:40px !important; }
  .pg-hdr-title{ font-size:18px !important; }
  .pg-hdr-sub{ font-size:11px !important; letter-spacing:1px !important; }
}
</style>""", unsafe_allow_html=True)
HERE = os.path.dirname(os.path.abspath(__file__))
LOGO = "https://chjk.co.kr/web/upload/category/logo/v2_c8bcd54017bc5f8880bb32d3de5333e6_BWrlZyel0_top.jpg"
LOGO_AUTH = "data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAuIAAACoCAYAAAClv4scAAAAGXRFWHRTb2Z0d2FyZQBBZG9iZSBJbWFnZVJlYWR5ccllPAAAAyFpVFh0WE1MOmNvbS5hZG9iZS54bXAAAAAAADw/eHBhY2tldCBiZWdpbj0i77u/IiBpZD0iVzVNME1wQ2VoaUh6cmVTek5UY3prYzlkIj8+IDx4OnhtcG1ldGEgeG1sbnM6eD0iYWRvYmU6bnM6bWV0YS8iIHg6eG1wdGs9IkFkb2JlIFhNUCBDb3JlIDUuNi1jMTQyIDc5LjE2MDkyNCwgMjAxNy8wNy8xMy0wMTowNjozOSAgICAgICAgIj4gPHJkZjpSREYgeG1sbnM6cmRmPSJodHRwOi8vd3d3LnczLm9yZy8xOTk5LzAyLzIyLXJkZi1zeW50YXgtbnMjIj4gPHJkZjpEZXNjcmlwdGlvbiByZGY6YWJvdXQ9IiIgeG1sbnM6eG1wPSJodHRwOi8vbnMuYWRvYmUuY29tL3hhcC8xLjAvIiB4bWxuczp4bXBNTT0iaHR0cDovL25zLmFkb2JlLmNvbS94YXAvMS4wL21tLyIgeG1sbnM6c3RSZWY9Imh0dHA6Ly9ucy5hZG9iZS5jb20veGFwLzEuMC9zVHlwZS9SZXNvdXJjZVJlZiMiIHhtcDpDcmVhdG9yVG9vbD0iQWRvYmUgUGhvdG9zaG9wIENDIChXaW5kb3dzKSIgeG1wTU06SW5zdGFuY2VJRD0ieG1wLmlpZDpFOEZGOTI4QjgzRDQxMUYxQjM0Q0E4MzdGOURBODgwQiIgeG1wTU06RG9jdW1lbnRJRD0ieG1wLmRpZDpFOEZGOTI4QzgzRDQxMUYxQjM0Q0E4MzdGOURBODgwQiI+IDx4bXBNTTpEZXJpdmVkRnJvbSBzdFJlZjppbnN0YW5jZUlEPSJ4bXAuaWlkOkU4RkY5Mjg5ODNENDExRjFCMzRDQTgzN0Y5REE4ODBCIiBzdFJlZjpkb2N1bWVudElEPSJ4bXAuZGlkOkU4RkY5MjhBODNENDExRjFCMzRDQTgzN0Y5REE4ODBCIi8+IDwvcmRmOkRlc2NyaXB0aW9uPiA8L3JkZjpSREY+IDwveDp4bXBtZXRhPiA8P3hwYWNrZXQgZW5kPSJyIj8+QGOD0AAANxBJREFUeNrsnQncVdPex1f1hCbKFIoyJK4pxRXaZIpLyJQx8z3G+yTe10zccA2X1MXluKaLKF6EXjNxlLFCIRQpQ8hQNKfe/+/Z6/E+93HO2fucs9Y+e+39+34+/3PO85x91l7z/q+1/uu/mijX8TKt5XVNlS4WqFz2W0UIIYQQQpylJgFpGCmyb8rKbbRIP1ZfQgghhBB3aep07L3MPilUwgkhhBBCCBXxqirhmM0fyiIkhBBCCCFUxKPlTJHNWISEEEIIIcRF3LQR9zLYnDmYxUcK1I+Vlb+Bt7lIa/3fX0QW1kkuO4+ZRAiJsE9qofukZiJtRJaLzNeyRPqkucwkQqiIu8QQkbYsvtQ/3DrJ6w4iPUQ21bK+ftAV+90CeZ0p8rnIJyJvibwp8rE8EJczYwkhZfRHGPhvpfujrUW6iGwssp5Iy4DfLpbXGbpf+kzkHZG3695z2aXMXEKSSxMHOzt0cJOU6xtNK2O0dM7p85riP+j2Fukrgo26nQzf4UeR/1W+V5qnJY9/Nhj3XoYHvsskfq8azt91lHlzrxkSzxklxKGzvHY2HIepEofZDuTVbInn1CrHwU7++Xt6elWlTtntk1BXDxLZS6S3SAvDd1iiJwker5Nc9qME9edYIdiyqm0oXDy7KfMTf29LPH+JMK8xOGxj+S4fROpW2c6zohpMdHFGfFjKlfD04WX+IK9niBwhsrrFO7UTOVrLQrnvg/J+s3QuEwyE/aTIagbjOtfCwwGDm7sMh3m5yGUlXH+8Mm92doLI3Q7k1T06/dWMQz6yIqdUGAZMxF6qUp0y3R+tKq8DtOxg+W4r6QEM5Fq5NxTxO+vKPZf9zvGeHWl6tMptKAw3iuxqOMxtlb/qEUV9XUVeXxZpZflOV4tcEGH9sfGsqAad3VJovczBetaBpEMB7ynyjHx6X+R0y0p4Y1poBe5ticN4kb1YICSl9Nf7LtLeH20gcrN8+lLkpgiU8Hx0FblG5AuJywiRbVk9SQB7RKCEgwOZ1SWzSGSWO4q4P6q7nuWWigfepiJPyKfXRPrEIEY7ijwrcRpbNzggJF1g5SW95zV4mbVFbpFP0/SEQOsYxAoz5UeKTJS4jdSrhoRUU0HeXOphF2Z3SXyCfWkuzYgPUsmwByKFH3griWDZebLy7cDjBpYnMTt+k16eJiQtHJvC/qiZyECFTdxKnaZ8L0xxpL/IexLXW0VWZ1UlDeowdLz9E6j0JwX0LY7YWnsZ7Dq/iGWW6A6jS52Sq9Slyp/tiSvY4Ax79Q8lzn1YcCQl7Cv1vV3K+iPY1cI+eDUHYgy3iLDjn6pNOAkB24usE+H9DmCWl8RUdxRxfxNAK5ZZYh96+ykssfpuv1wBg8OnJe5X1M2cEZJsMDjun5L+6FDle+ba2cHYryXyP5KGu0Ras9qmnqhnqHfW3nBIOD5xQxH3MtgQM4DlldiHHpZ+4ZrLxYcGZsexUnMWC5KkgGMS3hc1EblKPj2k3J/4OV7B7SH22xAq4tEBnbIvsz00DsyIo2P03RWSZD74zlf+0q/L7iifE/kHC5OkgF7ad28S+yKsasEt4AUJStXmIiP0c5Sk7/m6ibxWYxMv7cTD44SNOGZgdmBZJbKTwCzy3xxPxbsih6lcdgkLlKSEoxOqhN+vzPufrjbTRfpJ/7SC1TaVVEsh7iNtqgWzP5A50jZ/jLci7mWwNHg1yyqRSjhsMF13RfmFyH7SkOayQEmKSJaZoD9bfJvI4QlUwntL//QFqywV8YhpqfyTZklxfjsBNs4z4hcqf0McSdaDbyvln37msjnKPAUvErnslyxQkjK6ShvePkHpwXPmJCrhJGHPWWyYrOZmY3pPCebjeCviXmZDeT2H5ZS4zgEbMh/SI2ZXWSpykDzkJrNASUpJhnmKl8GmsiFUwkkC6Vtl/W5/7cOcFOaj+g81MY3gdSI8Ujl54GjmrhHd6wfl23Cjss/TCjQGAlhlwSaWrcqs/yfJQ+5FFiVJMUfKQ/YcaQe/OqyEd1D+ylwUGxnRF70g8qrC+QNKzRT5RfnHW7fR/dJGIvBdjj1Ru4mU4wKOSjipp9obJtcWwSnU41kUBfk4voq4l+ktr4ewjBKGl9lF+cdD22SxyH0id4m8XlRR8DKYlcfSHWxDDxMJc1LmxRLmvSxMknLwkMVhVk852hc10X2E7VMox4kMVXDPmssuLXDN9/p9SoP4YfNoL+WbzByhwp3oSSWc1NefFrp9xmEwQEXcOUXc74DorjB5HQPK9SbLdxkt8hd5EM0KdXUuu0D5rgefk/gNkvcTReBOsdApZLfLb65kYRJSxwBnFXGljlN2N5N9rXD6bi77aFm/9icQXq4TL3OevENOU4VPHKYSThqyh6rc/PNzkU4VhnGArrsucYK0o7ujvmncbHhOFtma7ShxQMndylLYeGgNlMbTL7QS/vsH388iGABieRh+hBc0uuIp/SAkhPgc6OTJjV6mnbxea/EOz9Y9w8pVwn/fN30tAlevOJhnFJVwEqptVo6JCdHNeKCUa4q4l2krr1ewSBKGl8EszqWWQl+uMLuVyw439NBbKHJ1XQfin/YJJirfV/ivLExCfgMzbgc7GG8MtNeyFPZIhU1quewc4yHnsp+LwIxud5FpVMJJgedt07o6WPlzFYdbmXDNy8N9nFLElbpMlbdBhcSbY0U6WgobG8but/DQmyWCDgQHSsFX+HwWIyG/w60j773Mugrma3Z4ri4/bB/ulcu+JK/b6OcllXDSGGz2bV9hGFP0+RivG4hPPxZJMPGwEfcymIE8g8WRuNE5NkX9l6XQMWNtdz+BDSWfkOSwh7Tx9aSdfOVIfNEXrWIh3BnKXzVbFkkq/P0tl7P6EUuK7/gG73tXGNaO0kesLXX2WxZNYeIyIz5UxdeVIikf7Py34a4Qrr9O59HNhFTEdAPPjyMdmRSAPfvJlkI/kSfskphg4iCdcY0U8krAZFxfFkvcFXEvs6+87sOiSCS2Hnx/56mWhFQMzBoqnalyxTwFhxCtaiHckdpchJBq61LYGLmZgZDqFXCYpiw3EB7txGOtiHsZ+EcdymJIZKeAA5ls2IfBXvsmZjAhFYMTjEdVGEY3aetbOJDWEyyEiQ3cl7AakZhgQuH9VgaWn9Z9ymWx8mziBOk+2rc5iaUi7m+coXubZILT4WzMQD0gHcT3zF5CKgabqB8yEE68Z8W9zAbK38RmmielL/qE1YgkSBEf1+hvE+Ypq6h4HDBERTxP54jT2QazCBKLLbswbqAkxFz/D9OUzypWxH23aXHlUEvh3sEqRGIy2IRLzh0NhPRagGJezUECFXELDFF2ZkxJPOhtIUz4532FWUuIMTYWebDCMDCzvkuM02jjFE1sznyK1YfEhL6G9LlxAYp5+fGL92A9hYq4l+kmr39m9id6dG7DbnSsymWXM4MJMaqIjzAQTjzNU/wDxXa1EPIzkbkrJCQYEzPO8IE/4T/+49uLzzYQtqkZeyriBoH/5ybM/sSynaVw6Z2AENOKeC47ReEQj8o4TJTeVWKYvm1FbGwUe5ZVh8RksIn6bcIGe4L0BYvz/N/UrDgP94mNIu5lDlPxXsYkldPNUrgTmbWEGGUT/V7p3guYGe4fw/TZmhR4k1WHxIQ+hgab40v8f6nQTrwA0R6i44/crmO2J54tLYQJk5TJzNrfaC3taazhMNdhtqaOzvodduJ/qzAsmKcU8sJSLTOO7hbCxMmWH7DqxJJ9LPSL3WKeZlMKbiGF29SGzS5SNpurXPZDVtNqKuL+EcOdmO2JZ2MLYX4tDXg+s/Y3mik7tq8kXXSse81lZ8hDEkvQldhx7ithrJHXvSh8EnuZaqTPhnvcjyU9v7LqxJL2WtKBvwHSlIeyQoo4VqJhsrKygXvg5E8q4lVTxL1MB3k9n1meCjayEOYMZishxlm17vh3//COERUq4nieHC5yS8L7oumsNiQmoL2uZSCcz6QPyL8pE3bjXmaiMrPZErP318Q4P++StN5lKezRkpf9qquIK3WtSEu2m8SP0Fc21DE0ZiYzlxArYFZ8qvJP2cRG+kr2Dh0TG0Xc3zy6noWQPy0jLveo6q0GnyUKwDus5onEtllKPeMMKeI9pS20l/r4DYsuakXcy6AAj3Iwf2A3+UYM4xXnGZk1LIXL0zQJscP6dYp4Lvut9NXPyee9KwhrRwkDnljy9VGYdW+dgL5oThm/WVNVz5SsLas4FfEKFXFTnlPgLQ+buv/FootSEfcyyPhhDubNlyIn0y65ZFa3FO4PzFpCrNCxwecHKlTEAWbFL8/z/6g3bNqyFS5nUuBjBRt6QszpVl2VuT0QYWbETXEAFfH/JAr3hceJbO9g3pxHJbwsbM1C/cisJcQKDc03HhNZZEARz8fciNNl6+TmchTxaaxmxDCmZsOxUlXcI5lvSvKpofvtJYMImilHpoh7mTaqcpdY1QDLMCNYPcqCBzUR4hZrN3jgQlkeU2F4m0jfv0Oe/y+MOF22zGCWlPGbqaxmJKaK+OshvQCZMk/B3o29WXxRKeJKXajc8028QmSgVMwVrB5l0cJSuL8wawmxQrtGfz9gIMwBef73U8TpsmV6uaCM33BGnJgDGx7NHRkfVsE2aZ7Cw30iUcS9DNxGne1gnvxblPC3WDXKZmVL4S5j1hISiSL+pMi8CsM8XJ4BzRv9L83mZfD6tIhVjRgCGx5NrT6HVbDHG4x/X+kfmrEYbSviSl0vspJj+YFZV/o6r4x5lsJtxawlJAJFHH6DlXqkwjDhJaRPo//9lNoc9ldYOStOTHGAoXBQL8N6hpuizK1MYy/ZzixGm4q4l9ldXvs5mB9XFnRqT8Ky3FK4zZm1hFghn3s7E+Ypxzb6O2rfwT9bCneVMn/3MasaMaBfYaPjXoZC+0B0nnADZN+O/DWDKTmAhWlLEfcysMtz0V3hZyJDWSUqZnHMHn6EkOK0y/O/F0S+rfhB62Uaei6JepLD1jH0VMRJNelj8HlYqt336wbTQTtxa4q4UhmRLR3Mi7P1kiypDFv+vtsxawmJqG35s1+jDCishzT4++uI0xW3Dd5UxEncFNhS7b5NbtiEd6U/sDhNK+JeBh36EAfz4UV58DzG6hBrRXx1Zi0hVijk6eh+A2E39Cn+RcTpsnUORLmTArQRJ5XqWNjg2NdgiKWammBG3KRHOc6KK/PunS53UGHCzM9ZrApGFfHlyvxqy7rM2v8AeTzZwmBnfWYt0WATF0z2NqwgjN1Eeeioclko4TMijr8tm/Q1y/wdfYnbB555ZhoOcxMVH2cBO1ZQ/xozR9plaas0OGfAy3wgn7YwFAfYicfprJlZyt5k4mf2FXF/ieF0BxvubVK5JitiBixpe5kvLSh0nZm5/8HPktfdjIboZY6X17uYtSnEy6wk9WlJo7a8Qv7/oHy6oIKQ4WLtKJFr9UMOEx/RuC3LZX+Q+MPc0LRL1dXLjM93Eh8cmLQaK5w1Hpd8Pt5w2xgrr7vGJH0mZ5DL3Xg5zqAivoPk77pSZl/HJH8vlbjcHfVNTc6ID42sgzU7er6UfZdxPreiiHuZJjxoiRArwBNDvhMjR1SoiINj6hTxXHaptGHMjHeKMF2zLdxvjQp+u4EKXi1E/7kqqyTJg0lvdBtKeyxH6dzCYBwwUIdP9GyaC9WMIu5lkJF9HEz/ZfJw+J5t2zjTRXoZDrOV8pfIP2X2EhIRuewU6d/hP7iSDfhbSRjbSFjvKn95NkpF/GsL99uggvwMPmfBy3CygeSrF7A62MRgiFuqeDjWODDtinhTA5UDh/bc4GDaYed0C1u3Fd6zFG53Zi0hkT8LTGzaPFq/fxZxuj6yEGYXVhdSJYU1iewhemRrdr6VMdDwKC0q4K6Qx6a7pYjvyKwlxArFTCEeNKKIexk8b2ZEnC4b+386SVp4wBiJmqQegIM9HH3SXLCVKeJepr28XuJgup8QJfwZtmtrTFRmXRzVsyezlpRBCwthpueAqVwWynOlJ+qtJ4ITlz+JOPbvWwgTe6E2YbMikeFl1lHY2JhcUu3GsNIZ8StF2jiW5qUi57BlW31ww/3PFAshb13nCo0kmaUWwrTheszGUuqSGJfLCANhYNNm1C78plgKdyc2VRIh2IfXJMHp66t9pFMRL3GEBnvdEx1M842iKH7Cdm2dly2FezSzNtHYOISltSNhLohxueCUzeUVhnGw8l0YRjkpAC8tMy2EvDObKomQpM8YwyVor7QWbnmKONzIKTXMwRHatyJXsE1HwlOWwj1B25qSZPKzhTDXthDmWhbC/DG2pZLLou98rsJQsHq6pyXFuBgvWQhzNzZVEgleBit6e3CwQUW8MYc7Onq5MJT7KGKCF5Sd2c2uisfiJpnvLITZ2UKYG1kI85uYl80DBsIYoOx4MinG81bqlJfpyeZKImBvlY49KVTESxidYePTtQ6mdZLiqYHRkcviRLvHLYU+ROphDTM5kdjwE2/D7eW2FsKMu8ncYyKLKgwD3hF+iDjeL1kK90g2V0IF1RgbyXN9yzQWcDkz4ucp86cmRkGtKIfL2aYjxdbAByd7DYokBVD4vUxXFmVk2PAz3V7KcDODdWKzujDNMzXWJZPL4nj2MRWGUmNpEFMs3l/K6xsWQj4u7f6PifXnDzYw9uWgg4p4w0oBBfxcB9M5UjrjV9mqIwfmKZ9bCvtKqY923Tl5GZgf5ETGy+cNWZyRKE0wZ7JhunCIwbAOtRC/r7TCGHdMmKdU40CcURbCXE3RAxexC0yAV6ciTkW8ITBJaeFYGhc6OnhIglKFFYgbLYWOAzUeEQXZvD9fbEb2MqfLp3dEeuqO8HHOfkXGeAthnqJPAa60biCMjIX4veJI2TwpUuk+m2ps8r9P2XGNeZ7UCZ60SaiYmmF7aU/rUREv/ACCu6YjHEzjtaIQzmR7rhr/Uva8QaDBjpW6ubVBJbybVopuVv/pIx+2a/drj0HELjaUUqzm/beBcM5SdkzznnKiZPy9H484OCkAry+PWggZE1MPSb/Qks2WWOCAFKZ5/7QlONyGN99d3HAH0zdLubmxNDnksr9I/blGPl1t6Q4dFE798zJny/vtZe8D8DIbKP+U2JNU4Rk7dIpDRC5mwVoFm3yXhe6fwjNYyvl1qSMvlFlHPF3+plmi7G1stgHMU453sF79XaS/hXC3UVgp8DL96JWLGMPLYC/UxoZDHSsy12B48Oayt+E4YhXgNirivwedbncH03eedIwL2KKrDgZxZ4rYOhUTs1G3ivxZOi8MvB6Rcl8WcoC5i1a+4ZKzeYh7XSS/e0/CH8VitTZ4+0Hy+AULHXxzrTCdobCROJddUcJD8ei6gZ5SK1lI8WiJy08OlRDKBjPMaztWr96ScnzGQr0C8CuO8E+U+4wrU/HCBEBvlQ5XdSScQmoSTFIdIPXT3FkN/jMUXpBWMxjPPerMQDGJFz0b6FXxqFgi6fygJkRGryqvf3OwEqMzfJBtORYPwIVSj+Dl5CHLd+qhsDFXqZ/kfs8q31PCx3V/+z7N0Vm0U74v8m7KPyRhzTLuc7eEP03SNZGFa43bLClMUHLuULDz9jIwP3qioBLsZWCatJ8I9gt4lgeqLrXnXyVvRunBtWtgNQsuFG2YmG0q8qrkDezob1HwX57LLg2hfGNG/U8ix4j8gU2fWFLEJxtVwv2+YHndKqPZvhqTHfuIPFyFPL9cS1TAmUXnMDPiFzk386EUZroGljTjRWw/vB+WBjtaRbP5pK3yl6D7WwofdqGPSXq2l3R9w8K1AurKR3rQZIMdtKyQcpwu79P1YA2zRtiU20nf2/YpruMd9eh0v5OKeC77tpT3ncpfBbNFXy2L5F4YrL8nMkf5JgHNdP1qrxV37D1Zg82dNBqgYf/T9sb7Glt9mB3zlIfTUtw1AZUBHikGOZiuu6XDncDWHDtOVb4XkvYJSAs27MFry25S15awaI0rTJhpuUo+3WP5TpiR3ERLNbjA0RLCahN8vrvo1vM8rSjb7oew+rKTFkJKYX9lftVmnKW42lDw96s7wyOMiWkCCJrtuV6Fs5uNEz87/HBLunI1W15ha5uUg5XwgL2VBWuNey0+POLAA9ImXnG0LWO18UFH4/69wn4SQuKLjZXj1ywOyk0/02FC6qWlsAsr4l4GdnQuus65guYCsX4IYqNXkg7BOEHaykAWrDVl7zRlx/9ztUEf5Xq9GeFw3XpCXq9jIyOxwz+vYg/Doc6WOv+ppbaEyc/JjgxGHFLEsSSg1FAH0wM7zxvZkmP/EEQZDUtQim6QNrMHC9ZKXUEHn7QVLsweHSNp+87xspkir1McTgHq1eNsZCRmYKOiae9M4y3H2Ub4KVfE/VkoF3dvn017XWfA3oOkmHXAVvZzFqnFgY7Ls6+/p1b6qecTkpb7HR5I/CqvR4q8xCZGYoQNSwTbirgNE8LORg/rc0oR9zI4zvtyB9PynHSsnN1w5yEIswO4hbvG8ZTA7m5nSc80FqrVunJCXRt3nwskPTcnqHQedLxu4ZyJvgmpW4VYxE7EEbxMM10fXVPEbdmfp+Jk0Xwz4jg5rp1j6cDMxiC2YgcVrFz2fK1kubiSAROb3bgnIZK6gvrRT2SMoymAOQpW7K5OWLnMsPgQjloZvzthrQZmQ7tL+l5nB+KOKm5B/1osMsFyG4L9+WwLIafCPKVpo9EYfJqe4mA6/ikV4X22YWcfhHgAwq3hVEdijFMF/yTxHiSymAUYqcJ0YF17d4uf6gYRuezQhJbMiATULZxwhwkBuFhd6HhqsPegVmRbSRPNbtzChuI5ISKTXRuz7tuJXtohXYq4P8PXzLE04HjVwWy/zj8IJyn/ZEyYqsTVdyhmNXHi4+YS36dZaFWpJ7+KwKQJpxDOdSDGUIS20V46ksoolRSXpLnsbXUKLE7IdA/MSGJluJOk4x9p8cFMRTyQqFzAjncoT2KqiHsZLPu66PnhUulwfmD7TcRDcIE2VcGRz6NjFrtX6kbnueyprG+xqCvYJIgN5SNjrBSdUNen5rIzE14WWCF6LkHpwYmuu4gMEJnhQIzfUv5JoZ3rPFLlsgvZQTiIvzHRxgFZ4yNKARXxMqnRFWBl5R/e4xqwgbuNLThxD/YPFJbyvQyOIIdifoCyf9R4IeDd4kqJ01gWTOzqyVfyeoTUE6yiXCRysDJ/Gl2pYL8AVhaHa1OatPCAMn/MdTXrFjYI3yd1a5RWyHH2weYxiiHMTx4SuV3i+g47g0Rga2NiVIr4ROXbo69sONze0g5XlXo+L9mKuFJniWzkYPwHcfkt0YoWTuw6SBrhBvJ+svJP5Yyini7QD7mbJA5vsyBiX09g1nSo1JOu8n688s1WOkYci5eVv9nvgZTuG3hM+d45VklY3YJt7R1St+6U9166fh2kquPQYIbIo1rGSdyWs/EnChszv9P0ilUUbWWxtBMo4zsaDhk+1eFbfVRyFXEvs468X+xg3EcnyBcvKd7AsbR/aZ14me7yvq/IXsrf4Gnq4IM5Is+KwJb3SbnnL4ZT0UmZna1dYSGnR2iFyiSLIqwnMCm4QOoIZsd3EtlTyx9Fmhu+2w/Kt718SteXWVVoGSbLa16FeT9X8n1tFc0eo+jd8fkz5Lk68TLY0OnpPmhn5duUtzZ8xxV1SpQ/mzm2TnwPNUlijIUBjY1NiX3V/09amuLnPP/bU5lf0Yt6onJPZf4wokJtHh6oXD/AsW4w3UQ6lV11Q3cN2IYPoZaaYrwMlKut9INwU5FNlD8TupbImiJtGj3Y5ukO8HuRT5TvpQUPuzfrPvsPW5K8eoJZ2s11XdlCBCss62pZSz9k2zT6FVyi/qgVbpibYDAIF10wh5vM+kIa1K+muv/p0aAPWl9LW5FWyl8paLhkP18rSTAx+UrLTN0vvVdXz9Jl2kRIammiOxLMqrhmEI8Zyy7SWc1mMZJAhT2XXcqMIAH1BP1hM5q7EfZFhJCoFfGN5RUb5FZyLP53a9+vhBBCCCGEOIVvzzdzwo+qUw8sze7sWPy3kXiPkfh/xaIkhBBCCCEu0dAl3BXKt4V0CczoD9dLyoQQQgghhDioiOey2MR2gYNpgKucI1mUhBBCCCHETUXc5x4RF/0mX6u8TEsWJyGEEEIIcVMR9w8IGOhgOjoo/wRGQgghhBBCnCC/bbWXwUERrpl7wOH7ZjKY+JzFSgghhBBC4k7TAv8/V2ShY2nBgQnXskgJIYQQQogL5D+OeOaEeapTD3y3m2Pp2ULi/ZLEn7PihBBCCCEk1jQt8t11yj9y1zWG6SOHCSGEEEIIcVARz2VhmnKug2nqJnISi5YQQgghhMSZ4INwvMwreHUsXd+JdJHBxFwWMSGEEEIIiSNhTDjgznCFY+laS+QSFi8hhBBCCIkrzQKvmDlhturUo6N86u5Y2raXeI+U+H/PYiaEEEIIIXEj7KbGi0TmOZa25iI3sIgJIYQQQkgcaRbqqpkT5qtOPZbJpz6OpW9TifcbEv9pLGpCCCGEEBInSnHzN1zkEwfTOFR5mRoWNSGEEEIIcVMRz2WXyOsgB9O4mciZLGpCCCGEEBInmpT8Cy/ztLzu7Vg6f1K+O8M5LHJCCCGEEBIHyjmBErPivzqWzrYiQ1jchBBCCCEkLjQr+RczJ8xRnXqsLp96OpbWbSXej0n8v2GxE0IIIYSQatO0zN9dLuKaf24MOoaxyAkhhBBCSFyU09KZOWGR6tTjZ/m0n2Pp7Szxnizx/5BFTwghhBBCqknTCn6bFZnsYJqvV15mFRY9IYQQQghxUxHPZbFhc6CDae6s3HTDSAghhBBCEkSTikPwMo/I60GOpXu+wqmbuexXrAKEkLQyfPjwFvJ2rshRIhuKLBR5W+T62tra/2UOEUKIXZoaCOMckcWOpbuVyN9Y/ISQFCvh6AdfErlMYWJCqeYiq4rsLjJGvv8v5hIhhNiliZFQvAyU2vMdTH9Plcu+wWqQWkVkJXlbT6SDyPoiU2tra99hzpCU1P8bVbB5YXdpE5OYW4QQYocaQ+FcKXKcyLqOpX+YDCJ2FGV8BatCKhSPXvL231r5huLdvtEl2DtARZykoS3U6D47iBNEqIgTQoglmhoJJZf9Rbk5I76DyDGsBqlhE5EDRLbLo4QTkibWUv6Jw0FsyqwihJC4K+I+94q86WAeXK28TCtWBUJIivgl5HXzmVWEEOKCIu6bd7jozhBmCheyKhBC0kJtbS0OZAtjcvIyc4sQQlxQxH1l/HV5vc/BfDhbeZkNWR0IISkCExDLi3z/ici/mE2EEGKPGgthwlb8YJGWDuUDTtq8TuRQVglCgtEeZ7DhdYNG8mZtbe3tzKH4I+X0tJTjAPl4i8hqjb5+S6S/XLOAOUUIIS4p4rnsl8rLXCWfrnAsLw6RePeW+I9ltSCkoAKOja5PiKxT4BLaFLuljI+QMh0jH/dV/qnD9Qf6jJPv6E2KEEKcU8R9rhc5WXfsLgF3ht1FGf+VVYOQvLQuooQTN5XxufL2AHOCEEKip6mVUHPZRcr31+waW+sBBCGEEEIIIQ4q4r4y/rByc8f9FcrLtGXVIIQQQgghbiriPnBnuNyxPFlTZDCrBiGEEEIIcVcRz2XfVW66vzpTeZnNWD0IIYQQQogtaiK4x8Uih6vfu8eKe74MFfkTq0g0DB8+HO4uuyv/+Pm1lX/89lIRuE/7TOQ9kYm1tbVLHEgL3GFuK7K9SHuRdjotixqkZYKkZXEEeYp49NDxQBtsovP0C5EPRN6SePzkaJ1potO3k0gHEdSN2Tp/X5N0LQ8ZDlwxbiPSTaSjzqdVdHl9JTJd+W4ZZ1UxrZvpdHYVaaN87zTf67S+JHFbGONyqs/fHrqckL/NdT38UuQjnb/fVXifGl0fthLZSPkbi1vq+8wRmaH8Q4ymyr24IZ8QEguaRHIXL3OWVmxdA/7QX4phvJaqXDYRbuLk4Qll9S8K7iODfc/jNMDHUJfkQTqpjHsdL293FblkkIR7YwVpwWZfmGP110pAMVB+T4rcKPd83XCe7iZvp4gcqBXKYkAhwV4OrFw9JHFZ1iCcHeTtgkbXw3Rr5yLhfSoyucj3h8k9luaJMwZf2QK/eVd+M7jR9bsr3ztTt3z1RK5fNUQ+9ZS3M0QOEFk1RNZCYbxH5HYJf06ZZbO/vJ1U4GuEO6bBtc3kDXV2kMgWRYLFcfX/FBmiT8wMG5dW8nZ/kUumSHgXV1AP/yhvp+l+NCh/4SrxDZE7EKdSBha6HE/V92kT4ic/iIzRZfki3TQSQqpJTUT3uVl3lF0dy59HYhqv0SL9HFfA2+nB2XEl/AwPWRxAMkB+D3drteUqRIbT0lr9v8vOsOZeUIKwUnS4/B717C+Slq8qjEdX3db2KOFnUPZ213KZhIE8fUZ/t65W5kthIy35+CafEq5pWeReXVSDfRsSx3Pl7eoiEwmTA/JpY3n7hyp9xQv5izMSLpIw/irvNzQcuIRk4yLpnK4VxPqyRB3fNkSYqH/wUnWQ/O5PEqdpIePSPKB825ZZD7GqMEwrxqVMCvXUcqmEcT58nAfcBwdK3aQHUqWwen0/IvKOhHOp3OsJqgOEkGrQNJK75LJ4+A5idhP9AP2D8k/uO66CYI4UmSBhbVXltEDpxIx2poL2dLBOy/YVxOMohFGiEt6YTUVw2uI/tTmBaT4p83frN0jnmfJ2jSq+mvdekXxCvZmkKjM7a6Xj8LyEt6bB/Omi44hB0dshlfCGbCLyjPy+bRXbQx8otyUq4fnK+34Ja5Qe5Oa7jydv75ahhDcGKyqPS3hPiLRn70wISaYi7ivjT8nrU8zy1CvhWGKHKcTGBoLDkeovSpibVyktSMMrqrjZQFjW0YpdjzLiUat8E4NWhpIGs5Y/WsiyaWX+ro2kcTU96Lo+xPWTC+QTlHjMsrYxlJ5dRV6ScNcyFN6mWsGEyVLrMsPAwPDaKrWHI3Ufv4ahIA8T2SfPfTBAeVr5ey9M0VcPrgghJKGKuA9mxZcx21OrhMNOFEvvJmcREdZjEnabiNMCUwosZ3cwGCzy5xEJe/US4gHTgmGGk3d1bW3tqzFSxOsHXTeIhJmpn5wnnw5SvjmKabYUeVTCb24gLCjR2APRosJwTpT4rBdxe9hR+TbXJp8p90g9fLjRfVZWvslOS8NJgN34+eylCSFRUxPp3XLZj5SX+YeimUpaQdl3KvL9Yq1UjhSZqnzzAyy376/rTCEFFSYVQ0TOijAtmD0rNhOPASfstTEDC+8ky7WitZ/I2cr3DFNI4bxOFd7Q11ApwRL+XSHjC/tjbDaEZ4p2Ol//kOe6cSKX6s8wKZubp88oNvMOzyWFNtp9UEF+wwRhz5DXTm6UT6hzd4b43RdaEUZeLdCDLJiJ9Ar4HTavXtIg38qleZ46jj0Q9yrf7GipVvxhBlXMjKKZzq9bI1LC22rlOMxgBJ5n3hf5Vvmz/mgT26jfmxqhrp6Z5/cwZ+sa4h7Is0k6/xA/rDjB/hyrGJ3z/AZ7I2aziyaEJFsR97lc5BiRtZj96UF7Nji2yCXfQNGSh+GURv+HHei78nsoUs8XUX7PlGuGye8/iyAtUGDPKHLJjzotExv9H2mbIr+H8owl/O4Ffo8ZTWwEfD8gKler4OV5KHBn5vPMIvdYRyt1A7UCiHgfWb8BUXvwaNvoN71VcU9C/5Tf2RgQnRPyull53DFep4I3HqJfujLPZtLB2mZ7RIDye55c9+8SNkqGAR5EzpIwf2nwP9hNY4D3egGFsh4vKkVc+S5qOwVcA8UaG5Kfy1MPUfdO0APpjnpA3r9RuusZEHAfzKAPkN8uyvPdrdrlJZRxOA/orwcAo+X6+9lLE0KqQdPI75jLYobtEmZ96rgs4PsBeZTw39AeRWAzWsj/bzMV3Yz4par4ZsFT8ijhDdOC2cBDtcJRiP8OGAxgk9lRAfGEsr9zIfeImAEUgfcPrChAETm+mr6yA2g84MDM+y3K33SJwRnsx/dW/mpDw3z6o643xThX0n1ZIY8u8v8XlT8bX8xlKExmLjSY3ivkvifnU0blfxi0Dgn4/aYRDbCxKnNmiMHgDvmUcJ2eH0Su1+WIVbOB8nehDbfFNjRjJeakAkp4/b1WiIwVOUL5m2Ef1Uo5IYRUhZoq3fd25fuX3YZFkHzkYb2hVpIK8Vyhh3Sjh+j7EhYUxkIz60fJ9+eU4VKulLS010p0IXAwyUMh0vKZnuU/rcAlh8r3pxXxp/yXgFtA2T86zKFBcg0OhjnGoSoF05H9JN4fNfp/voHc6QFhjRf5e4g8wkrG4IBrj5Rr4It+boXpu1vCCJqseL7EgYstYEK1coByfEiYPNGDjtoiba9twL2+kjDmhY24XIvVtoMVIYRUkaZVuWsuC3vZgcz+1BC0nHx3CWGNLPIdNm72tJwWeIZoVuT7O0sI68Ei38EOu3cBhQTf9Q8I+xpRNH5MYF3CwGL/PEp4JfkU9kAXmHoUOzAHhycdVGH6phYZnDVUImcq/xCcQqxuuyC0mUfQXgaYKn1u6JaLAr5fv5quGwkhxB1F3FfG4cLuYRZBKtg34PvnSwgr6BTKXSynZZ+A718oIaw3lb+JsxC7Fvl/Mfd2UNCSavP6V1HsPgx5LfKpmAcSbMh8JuyN5b7zQ9TVPhWmL1vMtKIR86rct8McqGPANfeaupnOl2KDS5gH3WTIgw0hhCRcEfeBHewiFkNy0bOS2xW5ZIa2mQ77MIabsWLXb20xLTDlKuZB44dSNutpxWJGGWnZPSDot7QdcdKA4nlTCdfvFvD92DCmO414NuD7HStMo0vHrQfVwy8lf98xfM+XA74/WgQmbEdrV4eEEEJFvCC5LJSQv7MYEg08jBQz5fi0TIWsEBtZTAtc/rWKMC2FDj0KOnFxQkLr0oOl2AArf8a2GOW4Uwyaje8ctU/7KlKNenhPiGtwQul9IrOlLP4lcoBIC0UIITGkJgZxgAu2E0XWY3EkkiDvDSu0S7xSKOaxpKPFtHQJ+H6Z4bR0KDMeHye0Lo0o8fogf9MflRGHMCseHUMo7ElgkyrUw9Eib6ni3lPqgb34SVoWStvEagb8xD+Wx8UlIYSkVBHPZecrL3Ou8mcwSPIIOnlyDy2maGkxLesHfI+Noi8ZvF+rMuPxRQLrETxqjC/xN0FnFcwpIx5hNsCunRJFfIOA7780fUNsrBWFGhumsVeklBN6MSN+oJabJQwM6q6X8KayiyaEVJOmMYnHCBW8CY+4ScsE3S/y5W1RGJo1+juM2cPCBNajdwr5+S5jIFPP/DIUwQUhLlstJW07qC7+YuOmUgZwX9lb+W4sy+0jTlb+4Vq3pMiUiBBCRbwAuSw2KNGdYTKJ2p3YAothR/3AXi5KR+MDjFYJ8bv5CaxHv5Y4gAlTVkvLjMsSJ/rV6ivi1vz561NncQ7F3yu4Dwa5cBX5tj4tlxBCUqqI+8o4XLndwyJJHFF7xbFp+7ks4rTMK1N5TP3GNFHUfg5xWbleNVZK4UAoH0ErL00tl/F8EXjewgbt60S+LzMo7GN5WZTxTuyuCSFRUxOz+Fwgcogq7iOZuMX8EN/PMXg/mzafQcodFJNvDd4vn613mOX+lqx2dSwOULZLHrCEnGmfn6K2Xcz8Z9WIBl2z5O1cKRs8P3qL9FO+Lfj6JQQDe3Oc+NyHzYYQkl5FPJf9WnmZq+TTVSyaxBCkmD4vD9J+jqQlyDf365KW3W1GQMKHZ5Z5AUrOOqx2dXwlsqHhfAqzQfCLlOQvBtBrV5hXJtsGzJde0PIXaSdwr3io8k/2DaOU7yW/2V7CeYtNhxASFXG0ZbxB5DMWTWKYEfD9Rg6lZWZM0vJJwPddWe3qCOpHupQRZpDLPphizUpJ/gbVw6raXYtCPUnkIvnYWSvkYfz8H8JmQwhJtyKey2I5+RwWTWKYHPD9ZsOHD2+dkLR0krSsHUE83g/4vgerXR3vBXy/ZRlhbhPw/btwsZeS/A1y0dg9DpGU8sCm5//R8QkaPHRjsyGEpFsR95XxR5VZf8ykeg9BmAcUm0luroKPIo9LWmAjPiXgsihsTIOO+e4Z0YAg7rwa8P0ukk+rlBhmUPk+n6L8DeqjMTDdKkbtd668DQu4bE02G0IIFXEfuDP8lUWUCJ4O+P60BKXllAji8GyIdn2ipXsHtcl2MSqrF1VxTzdQwvcNG5golTggaNeAy55K2UAnyHPKyTGLc9CMOH2KE0KoiNeRy8IMIMsiSgT3B3z/J1FyvFIDld80EdncYDybGEhLL4nTvuXcXH63RZjramtrsRkwaOb1/FLcscm1bUXCbN4Ocg+5QVwqneQTTsEcE3DZhZLusP0gTgAu5rrwQ7nnuLQ0an240aiAy07VmybD1sMWIq0CrllJ5FC0/zKiHTRQ/F4RQggV8d+4RNn1C00iGlaJvBNwzSh5sG5WwgMbCjhmPN+Rz2HtOoNmcwO9aIjygXQEnQL77xLihLRsLIKZ1HflfaeQP7s54Huc7vishLdhwL1rRDCL/7Fub0EEeY7ZScJrH6O6d2PA97CnvyxEGe0pb4MCLhuWwrZ9S8D3GLg8Ifm3dUD+NhU5XD5+VKzMtPKNCZqHRJ4rsc/AbzMBl72vCCGEini9+pbF7MRgFpPb6M1rl4ZQgt+Uh+XZxTZvync7iGAWDrbavfWD/sGgWTTNDwHfY5ZtdX2fVph5K3DdxQHhrCEyHn6NA9LSXeQ+5fs+30f5J/2NwOx0iLSMFnkt4JpNtXL/NyhC9TPe8t4Mtrsi5ynfk8StIjC7uEj+t2NAWcId5XcBiteohgMAKOYiZ4i8L9I54ro3Vg/Yig74JV4358v3BgOVx3X5FAIDmTtS2LZxENvDAZd10G17qEjP+nalle9NRWrlzw/QjpXvZvBk+V8hl6ZXiBynP++h/GPq79XhNinS1mD7jbYW5F50jCKEkAipcSCOmHE5VWRzFpfTD2zMij0hH/cvchnsM68X+atcC/tTzI5hRWRNraj3FFkvz+/grg+zaH8OiMb0gO/hfnC23Psbfb/DRB7Lk5YXtAJ9TJGwcFjMVVrJG6eV7e91WjBj/EeV34wD5iS3iRweNLiRcM9U/ux884A8PV8LFBJsWFutwLVQNO+Ta7aR8IsdHAQb9aOLfL8LFHwJ5ys92G+40nBlwG9tcIbyV2SKHe5zusixEmf4oIbbQ3hv6qgHex0Cwl8ucgp8vKe0eeN0S2xiLebbHnl/lpb6etimyGTQ7XLNG5KnXzdQprGX5MI8dfYYLV/INVh9w6x2/coNNi131wPdoMH6DJEn2VsTQqKkaexjmMvi4TaIRZUIThD5PMR1eGDuDX1T+TPpUJIOLqCE14NZtP4ByiuU4a8D7t1cK2A1AYOG0/VAIQgo5DBrgNI8WCuFh6rittT9JS0nhRjcTKxXbEpgtRCDkSATi7tD3ms99Xtzn6MkbdtFPAicqgfzQWD14kCdp+fpAUOHEL8brGfe0zrIhgJ7XBn1sNjzBwPWO+tnueUdbfGmgDDRbo9U/qz57Vow8DskhBIOTk3xYIoQQkW8qDL+DGcqEvHA/l4rpV9aCB7HbYfZvHVrCWHuV2gjn3ZliLRMt5AWHAqzLGSeYsXoasP3P1HSvWuRe2KjaCXuRa+tQt3D4OECC0H/Uyt7aW/bj+nBqUkwi32U/vy2FlucK2l4hr00IYSKeGHOFlnKInP+gT1N3nqJTDQYLGyAu0nYI0NcO1wFn5BZD0xItiuSFngvgbeX1wymZbxIdwn7nhLyFAom7JgXGrj/EuXbwOcCrsPqxrdl3mM3UfT7VqHuYcBysk6jiXw6W8I8PUUH+ATlLwYlWO350UBwyFOcsvyQDvtr3W9cbqj86oEJ1gAJ/zqWICGEingxcln4fx3GIkvEA3uGvGFTII6fnldBUJNEDhLZUyv4Ye4Nm3Mogd+EvMf+AeFBQYBN9DkVKiCwa8Wyei8J88My8hSeJOAmrpJZvdF6QHMlTiMMuN/nehDyYRn3GaOq5J1C4o0NlfBoU8nBOzhQaQcJayhb8+/yFydYwkPKSK1Ml5u/PSWsc0SWNAh7qchl8rGLyJ3KXzmqRNEfIbKFhHkfS44QUi1qHIvvEJFjlb8Bh7j9wMYD9ip4q5D3ASL9tXIeVCdnKd9MaYSE8WqZ956s3an9Vfkzu4W8o0C5nxEiPJiR3CBhQhk+Wqellyrucxp8rdPygMjYSmdW5fewWd9H4rG9biewjV034GdQqHGS7R3y+ykl3u9j7SMam+hO1wpSIeYo3/PIbdrTRmPmBQy0pxmsexg87AUPPLr8YRce5LoSHndgfvFv+f3LFdx+UkA6J5UQFmagWxT4LszqyGIbea5Xio6Q/EX7wl4HeEDZKERbQP24s0D9aBg+VrROkvDP1W0Ng3G4/QyyA0f7elcEm8bvknA+Y09MCKk2TZyLsZfB0vLtKS+30SqX7Ze0ROEwD3nbSj+019GK7BKtVOCh+b48PL80fM+W+iGOe66m7wdlf1IlD2oJd+UGaVlXp2WZTssMnZZZEeQp3MF104NXeLXAKtgCrYC/p5UmU/faWN62VL5JD+4FO/rv9H2mxbzuoZzgkxpea3DiJrxxwLMH9jW8JzKdJigV5S/aAAa/6+l2VqPbwizdFqZXGD7KC561Oit/0+Yq+h7zdDmi35iiV8QIIYSKeAWKOBQJbNrZNsXllkhFnBBCCCEkTTR1Lsa5LGxXB7LoCCGEEEIIFfHolXF4dBjJ4iOEEEIIIVTEowcbdRaxCAkhhBBCCBXxKMllsXP+GhYhIYQQQgihIh49OKHvCxYjIYQQQghxjRqnY5/LLlBe5hj5tFvKym0qqy4hhBBCiNv8nwADADF0DsSo+esGAAAAAElFTkSuQmCC"
NAVY = "#1B3A6B"
GIT_TOKEN = st.secrets.get("github_token", os.environ.get("GITHUB_TOKEN", ""))
DATA_REPO = st.secrets.get("github_data_repo", st.secrets.get("github_repo", os.environ.get("GITHUB_REPO", "")))

@st.cache_resource(show_spinner=False)
def _data_repo_dir():
    return store.clone_data_repo(DATA_REPO, GIT_TOKEN, os.path.join(tempfile.gettempdir(), "chjk_data_repo"))

_REPO_DIR = _data_repo_dir() or HERE   # 자료 저장소(비공개). 실패 시 코드폴더(번들 data.zip) 대체
DATA_OK = bool(_data_repo_dir()) or os.path.exists(os.path.join(HERE, "data.zip"))
DATA_ZIP = os.path.join(_REPO_DIR, "data.zip")
BACKUP_ZIP = os.path.join(_REPO_DIR, "_직전본.zip")
BASIS = os.path.join(_REPO_DIR, "_기준일.txt")
DATA = store.ensure_data(DATA_ZIP, os.path.join(tempfile.gettempdir(), "misu_work"))

def basis_date():
    try: return datetime.date.fromisoformat(open(BASIS, encoding="utf-8").read().strip())
    except Exception: return None

def header(full=True, title=None):
    title = title or ("미수관리 시스템" if full else "관리 시스템")
    bd = basis_date()
    sub = f"현재 자료: {bd.year}년 {bd.month}월 {bd.day}일 기준" if (full and bd) else ""
    st.markdown(
        f"<div class='pg-hdr' style='display:flex;align-items:center;gap:18px;border-bottom:2px solid {NAVY};padding-bottom:12px;margin:6px 0 16px;'>"
        f"<img class='pg-hdr-logo' src='{LOGO}' style='height:52px;width:auto;object-fit:contain;' onerror=\"this.style.display='none'\">"
        f"<div><div class='pg-hdr-title' style='font-size:22px;font-weight:600;color:{NAVY};line-height:1.25;'>{title}</div>"
        f"<div class='pg-hdr-sub' style='font-size:12px;letter-spacing:2px;color:#888;'>{sub}</div></div></div>",
        unsafe_allow_html=True)

# ── 업무관리 앱과 로그인 공유 (동일 Firebase 계정) ──
_FB_API_KEY = "AIzaSyBTYey9GzRIRRqiiZoQ3gpxI-Ty1BhXyZU"
_FB_DB_URL  = "https://chjk-scheduler-default-rtdb.asia-southeast1.firebasedatabase.app"
_FB_PATH    = "teamdata_test"
MISU_ADMIN_IDS = ["chjk", "김소준"]   # 자료처리 전체 권한 (그 외 계정은 조회 전용)

def _sha256(v):
    return hashlib.sha256(str(v).encode("utf-8")).hexdigest()

def _synth_email(uid):
    return "u" + _sha256(uid)[:32] + "@chjk-scheduler.web.app"

def _fb_login(uid, pw):
    """업무관리와 동일한 Firebase 계정으로 검증. return (ok, role, err)."""
    if not uid or not pw:
        return False, None, "성명과 비밀번호를 입력하세요."
    try:
        r = requests.post(
            "https://identitytoolkit.googleapis.com/v1/accounts:signInWithPassword?key=" + _FB_API_KEY,
            json={"email": _synth_email(uid), "password": _sha256(pw), "returnSecureToken": True},
            timeout=15)
    except Exception:
        return False, None, "서버 연결에 실패했습니다. 잠시 후 다시 시도하세요."
    if r.status_code != 200:
        return False, None, "성명 또는 비밀번호가 올바르지 않습니다."
    d = r.json(); token = d.get("idToken"); fuid = d.get("localId")
    try:
        a = requests.get(_FB_DB_URL + "/" + _FB_PATH + "_authorized/" + str(fuid) + ".json?auth=" + str(token), timeout=15)
        allowed = (a.status_code == 200 and a.json() is True)
    except Exception:
        allowed = False
    if not allowed:
        return False, None, "접근이 허용되지 않은 계정입니다. 관리자에게 문의하세요."
    import time as _t
    st.session_state["_fb_tok"] = token
    st.session_state["_fb_ref"] = d.get("refreshToken")
    st.session_state["_fb_tok_at"] = _t.time()
    role = "admin" if uid in MISU_ADMIN_IDS else "view"
    try:
        import urllib.parse as _up2
        requests.delete(_FB_DB_URL + "/" + _FB_PATH + "_logout/" + _up2.quote(uid, safe="") + ".json?auth=" + str(token), timeout=10)
    except Exception:
        pass
    return True, role, ""

def _fb_id_token():
    """저장된 토큰 재사용, 만료 임박 시 refresh_token으로 갱신. 실패 시 None."""
    import time as _t
    tok = st.session_state.get("_fb_tok"); at = st.session_state.get("_fb_tok_at", 0)
    if tok and (_t.time() - at) < 3300:
        return tok
    ref = st.session_state.get("_fb_ref")
    if not ref: return None
    try:
        r = requests.post("https://securetoken.googleapis.com/v1/token?key=" + _FB_API_KEY,
                          data={"grant_type": "refresh_token", "refresh_token": ref}, timeout=15)
        if r.status_code == 200:
            j = r.json(); nt = j.get("id_token")
            st.session_state["_fb_tok"] = nt
            st.session_state["_fb_ref"] = j.get("refresh_token", ref)
            st.session_state["_fb_tok_at"] = _t.time()
            return nt
    except Exception:
        return None
    return None

def _publish_misu_summary(summary):
    """미수 요약을 Firebase 전용 키에 저장(업무관리 카드용). 실패해도 화면에 영향 없음."""
    try:
        import json as _json
        sig = _json.dumps({k: v for k, v in summary.items() if k != "updatedAt"},
                          ensure_ascii=False, sort_keys=True)
        if st.session_state.get("_misu_pub_sig") == sig: return
        tok = _fb_id_token()
        if not tok: return
        r = requests.put(_FB_DB_URL + "/" + _FB_PATH + "_misu_summary.json?auth=" + tok,
                         json=summary, timeout=10)
        if r.status_code == 200:
            st.session_state["_misu_pub_sig"] = sig
    except Exception:
        pass

_LOGIN_CSS = """
<style>
@font-face{font-family:'Pretendard';font-weight:400;font-style:normal;src:url('https://cdn.jsdelivr.net/gh/orioncactus/pretendard@v1.3.9/packages/pretendard/dist/web/static/woff2/Pretendard-Regular.woff2') format('woff2');font-display:swap;}
@font-face{font-family:'Pretendard';font-weight:600;font-style:normal;src:url('https://cdn.jsdelivr.net/gh/orioncactus/pretendard@v1.3.9/packages/pretendard/dist/web/static/woff2/Pretendard-SemiBold.woff2') format('woff2');font-display:swap;}
@font-face{font-family:'Pretendard';font-weight:700;font-style:normal;src:url('https://cdn.jsdelivr.net/gh/orioncactus/pretendard@v1.3.9/packages/pretendard/dist/web/static/woff2/Pretendard-Bold.woff2') format('woff2');font-display:swap;}
[data-testid='stSidebar'],[data-testid='stSidebarCollapsedControl'],[data-testid='stSidebarCollapseButton'],[data-testid='stToolbar'],[data-testid='stHeader'],#MainMenu,footer{display:none!important;}
/* 세로 중앙 + 카드 420px (컨테이너 452 = 420 + 좌우여백 16) */
[data-testid='stMainBlockContainer'],.stMainBlockContainer,.block-container{max-width:452px!important;padding:16px!important;margin:0 auto!important;min-height:100vh!important;display:flex!important;flex-direction:column!important;justify-content:center!important;}
[data-testid='stMainBlockContainer'] [data-testid='stVerticalBlock']{flex:0 0 auto!important;}
/* 카드 */
@keyframes _misuFadeIn{to{opacity:1;}}
[data-testid='stForm']{opacity:0;animation:_misuFadeIn .2s ease .45s forwards;}
[data-testid='stForm']{border:1px solid #c8d2de!important;border-radius:0!important;box-shadow:0 14px 40px rgba(15,23,42,.16)!important;background:#fff!important;padding:0 20px 20px!important;}
[data-testid='stForm'],[data-testid='stForm'] *{font-family:'Pretendard',-apple-system,BlinkMacSystemFont,'Segoe UI','Malgun Gothic','Apple SD Gothic Neo','Noto Sans KR',sans-serif!important;}
[data-testid='stIconMaterial'],span[data-testid='stIconMaterial'],button[aria-label='Show password'] span,button[aria-label='Hide password'] span{font-family:'Material Symbols Rounded','Material Symbols Outlined','Material Icons Rounded','Material Icons'!important;}
/* 간격: Streamlit gap을 끄고 요소별 명시 마진으로 고정 (업무관리 실측값) */
[data-testid='stForm'] [data-testid='stVerticalBlock']{gap:0!important;row-gap:0!important;}
[data-testid='stForm'] [data-testid='stElementContainer']{margin:0!important;min-height:0!important;}
[data-testid='stForm'] [data-testid='stMarkdown'],[data-testid='stForm'] [data-testid='stMarkdown']>div,[data-testid='stForm'] [data-testid='stMarkdownContainer']{margin:0!important;}
[data-testid='stForm'] [data-testid='stElementContainer']:has(.misu-auth-head){margin:0 0 18px!important;}
[data-testid='stForm'] [data-testid='stElementContainer']:has(.misu-auth-err){margin:0!important;}
[data-testid='stForm'] [data-testid='stElementContainer']:has([data-testid='stTextInput']){margin:0 0 14px!important;}
[data-testid='stForm'] [data-testid='stElementContainer']:has(.misu-auth-foot){margin:14px 0 0!important;}
/* 머리띠 */
.misu-auth-head{display:flex;align-items:center;gap:12px;padding:16px 20px;background:#f4f6f9;border-bottom:1px solid #c8d2de;margin:0 -20px 0;}
.misu-auth-head img{height:32px;width:auto;display:block;}
.misu-auth-foot{text-align:center;font-size:11px;color:#9ca3af;margin:0;line-height:13px;}
.misu-auth-err{color:#dc2626;font-size:13px;line-height:1.4;margin:0;}
/* 라벨: 12.5px/600, 줄높이 15px(실측), 아래 6px */
[data-testid='stForm'] [data-testid='stWidgetLabel']{margin:0 0 6px!important;min-height:0!important;}
[data-testid='stForm'] [data-testid='stWidgetLabel'],[data-testid='stForm'] [data-testid='stWidgetLabel'] *{font-size:12.5px!important;font-weight:600!important;color:#374151!important;line-height:15px!important;}
/* 입력창: 총높이 42px, 안여백 10x12 */
[data-testid='stForm'] [data-testid='stTextInputRootElement']{background:#fff!important;border:1px solid #c8d2de!important;border-radius:0!important;}
[data-testid='stForm'] [data-testid='stTextInputRootElement']:focus-within{border-color:#14305c!important;background:#fff!important;box-shadow:inset 0 0 0 1px #14305c!important;}
[data-testid='stForm'] [data-testid='stTextInput'] input{background:transparent!important;border:none!important;font-size:14px!important;padding:10px 12px!important;height:42px!important;box-sizing:border-box!important;line-height:normal!important;color:#111827!important;-webkit-text-fill-color:#111827!important;}
[data-testid='stForm'] [data-testid='stTextInput'] input::placeholder{color:#9ca3af!important;}
[data-testid='stForm'] [data-testid='stTextInput'] button{background:transparent!important;border:none!important;box-shadow:none!important;color:#8a94a6!important;}
[data-testid='stForm'] [data-testid='stTextInput'] button span{font-size:19px!important;}
[data-testid='stForm'] [data-testid='InputInstructions']{display:none!important;}
/* 버튼: 총높이 40px = 테두리2 + 패딩22 + 글자16 (실측) */
[data-testid='stFormSubmitButton']{width:100%!important;margin:0!important;}
[data-testid='stFormSubmitButton'] button{width:100%!important;background:#14305c!important;border:1px solid #14305c!important;border-radius:0!important;padding:11px!important;min-height:auto!important;box-shadow:none!important;}
[data-testid='stFormSubmitButton'] button:hover{background:#1B3A6B!important;border-color:#1B3A6B!important;}
[data-testid='stFormSubmitButton'] button:focus{box-shadow:none!important;}
[data-testid='stFormSubmitButton'] button p,[data-testid='stFormSubmitButton'] button div{color:#fff!important;font-weight:700!important;font-size:14px!important;line-height:16px!important;}
</style>
"""

def _restore_session(rt, nm, at_ms=None):
    """refresh token으로 세션 복원(쿠키/SSO 공용). 토큰교환 1회 + 권한확인 1회(+쿠키 복원 시 로그아웃 대조 1회)."""
    if not rt or not nm: return False
    try:
        r = requests.post("https://securetoken.googleapis.com/v1/token?key=" + _FB_API_KEY,
                          data={"grant_type": "refresh_token", "refresh_token": rt}, timeout=15)
        if r.status_code != 200: return False
        j = r.json(); tok = j.get("id_token"); fuid = j.get("user_id")
        a = requests.get(_FB_DB_URL + "/" + _FB_PATH + "_authorized/" + str(fuid) + ".json?auth=" + str(tok), timeout=15)
        if not (a.status_code == 200 and a.json() is True): return False
        import time as _t
        # 로그아웃 표식이 있으면 복원 거부 (표식은 로그인 성공 시 삭제됨 — 시계 차이와 무관)
        try:
            import urllib.parse as _up
            lo = requests.get(_FB_DB_URL + "/" + _FB_PATH + "_logout/" + _up.quote(nm, safe="") + ".json?auth=" + str(tok), timeout=15)
            if lo.status_code == 200 and isinstance(lo.json(), (int, float)):
                return False
        except Exception:
            pass
        st.session_state["_fb_tok"] = tok
        st.session_state["_fb_ref"] = j.get("refresh_token", rt)
        st.session_state["_fb_tok_at"] = _t.time()
        st.session_state["_login_at_ms"] = int(at_ms if at_ms is not None else _t.time() * 1000)
        st.session_state.update(auth=True, role=("admin" if nm in MISU_ADMIN_IDS else "view"), uid=nm)
        return True
    except Exception:
        return False

def _try_sso_login():
    """업무관리 시스템에서 '입금매칭 앱 열기'로 넘어온 경우: URL의 토큰으로 즉시 로그인."""
    try:
        rt = st.query_params.get("sso"); nm = st.query_params.get("u")
        at = st.query_params.get("at")
    except Exception:
        return False
    if not rt or not nm: return False
    try: at_ms = int(at) if at else None
    except Exception: at_ms = None
    ok = _restore_session(rt, str(nm).strip(), at_ms=at_ms)
    try:
        del st.query_params["sso"]
        del st.query_params["u"]
    except Exception:
        pass
    try: del st.query_params["at"]
    except Exception: pass
    return ok

def _try_cookie_login():
    """새로고침 시 쿠키의 refresh token으로 로그인 자동 복원."""
    if st.session_state.get("_ck_tried"): return False
    st.session_state["_ck_tried"] = True
    try:
        import urllib.parse
        ck = getattr(st, "context", None)
        ck = ck.cookies if ck is not None else {}
        rt = ck.get("misu_rt"); nm = urllib.parse.unquote(ck.get("misu_uid") or "")
        at = int(ck.get("misu_at") or 0) or None
    except Exception:
        return False
    return _restore_session(rt, nm, at_ms=at)

def check_pw():
    if st.session_state.get("auth"): return True
    if _try_sso_login(): return True
    if _try_cookie_login(): return True
    box = st.empty()
    with box.container():
        st.markdown(_LOGIN_CSS, unsafe_allow_html=True)
        with st.form("login_form"):
            st.markdown("<div class='misu-auth-head'><img src='" + LOGO_AUTH + "' alt='충해전기(주)'></div>", unsafe_allow_html=True)
            _lerr = st.session_state.pop("_login_err", None)
            if _lerr:
                st.markdown("<div class='misu-auth-err'>" + _lerr + "</div>", unsafe_allow_html=True)
            uid = st.text_input("성명", placeholder="성명")
            pw = st.text_input("비밀번호", type="password", placeholder="비밀번호")
            submitted = st.form_submit_button("로그인", use_container_width=True)
            st.markdown("<div class='misu-auth-foot'>충해전기(주) 내부 운영 시스템 · 관계자 외 접근 금지</div>", unsafe_allow_html=True)
    _components.html("""
<script>
(function(){
  var tries=0;
  var t=setInterval(function(){
    tries++;
    try{
      var d=window.parent.document;
      // 자동 로그인 복원: localStorage 토큰 → SSO 주소로 이동 (탭 세션당 1회만 시도)
      try{
        var L=window.parent.localStorage, S=window.parent.sessionStorage;
        var _rt=L.getItem('misu_rt'), _u=L.getItem('misu_uid'), _at=L.getItem('misu_at');
        if(_rt && _u && !S.getItem('misu_auto_tried')){
          S.setItem('misu_auto_tried','1');
          var _loc=window.parent.location;
          var _u2=_loc.origin+_loc.pathname+'?sso='+encodeURIComponent(_rt)+'&u='+encodeURIComponent(_u)+(_at?('&at='+encodeURIComponent(_at)):'');
          // Streamlit 컴포넌트 iframe은 sandbox라 직접 이동 불가 → 부모 문서에 스크립트 주입(같은 출처라 허용)
          try{
            var _pd=window.parent.document;
            var _ov=_pd.createElement('div');
            _ov.style.cssText='position:fixed;inset:0;z-index:2147483600;background:#fff;display:flex;align-items:center;justify-content:center;font-family:Pretendard,sans-serif;font-size:14px;color:#1B3A6B;font-weight:600';
            _ov.textContent='자동 로그인 중\u2026';
            _pd.body.appendChild(_ov);
            var _sc=_pd.createElement('script');
            _sc.textContent='location.replace('+JSON.stringify(_u2)+');';
            (_pd.head||_pd.body).appendChild(_sc);
          }catch(_ne){ try{ _loc.replace(_u2); }catch(_n2){} }
          return;
        }
      }catch(_ae){}
      var last=window.parent.localStorage.getItem('misu_last_user')||'';
      var inputs=d.querySelectorAll("[data-testid='stForm'] input");
      if(inputs.length>=2){
        var nameInput=null,pwInput=null;
        inputs.forEach(function(i){ if(i.type==='password'){ if(!pwInput)pwInput=i;} else if(!nameInput){nameInput=i;} });
        if(nameInput){
          if(last && !nameInput.value){
            var setter=Object.getOwnPropertyDescriptor(window.parent.HTMLInputElement.prototype,'value').set;
            setter.call(nameInput,last);
            nameInput.dispatchEvent(new Event('input',{bubbles:true}));
            if(pwInput) pwInput.focus();
          }
          clearInterval(t);
        }
      }
      if(tries>25) clearInterval(t);
    }catch(e){ clearInterval(t); }
  },100);
})();
</script>
""", height=0)
    if submitted:
        ok, role, err = _fb_login((uid or "").strip(), pw or "")
        if ok:
            box.empty()
            import time as _t
            st.session_state["_login_at_ms"] = int(_t.time() * 1000)
            st.session_state.update(auth=True, role=role, uid=(uid or "").strip())
            st.rerun()
        else:
            st.session_state["_login_err"] = err or "로그인에 실패했습니다."
            st.rerun()
    return False

def zip_bytes(d, dirs_only=None):
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        for root, _, files in os.walk(d):
            for fn in files:
                fp = os.path.join(root, fn); rel = os.path.relpath(fp, d).replace("\\", "/")
                if dirs_only and not any(rel.startswith(p + "/") for p in dirs_only): continue
                z.write(fp, rel)
    return buf.getvalue()

def zip_backup_customers():
    """직전본.zip에서 서울/화성 거래처 파일만 추려 재압축(지원표 제외)."""
    buf = io.BytesIO()
    with zipfile.ZipFile(BACKUP_ZIP) as src, zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        for n in src.namelist():
            nn = n.replace("\\", "/")
            if nn.startswith("서울/") or nn.startswith("화성/"): z.writestr(n, src.read(n))
    return buf.getvalue()

def all_company_files():
    out = []
    for loc in ["서울", "화성"]:
        for f in sorted(glob.glob(os.path.join(DATA, loc, "*.xlsx"))):
            n = os.path.basename(f)
            if n.startswith("_") or n.startswith("~$"): continue
            out.append((loc, f, n))
    return out

def _cust_label(loc, n):
    m = re.search(r"\((\d{3}-\d{2}-\d{5})\)", n)
    biz = m.group(1) if m else ""
    nm = re.sub(r"^[^)]*\)\s*", "", n).rsplit(" (", 1)[0]
    return f"[{loc}] {nm}" + (f" ({biz})" if biz else "")

def _duerules():
    duer = {}
    try:
        wb = load_workbook(os.path.join(DATA, "_거래처기준설정표.xlsx"), data_only=True)
        ws = wb["기준설정"] if "기준설정" in wb.sheetnames else wb.active
        for r in range(2, ws.max_row + 1):
            b = ws.cell(r, 3).value; ru = ws.cell(r, 4).value
            if b and ru: duer[str(b).strip()] = str(ru).strip()
        wb.close()
    except Exception:
        pass
    return duer

def render_result(R):
    st.divider()
    st.write("**파일 판별 결과**")
    for loc, d in R["detected"].items():
        line = " / ".join(f"{k} {len(v)}건" for k, v in d.items() if v)
        if line: st.write(f"- {loc}: {line}")
    # 상태 카드
    nc = R["new_companies"]; stt = R["status"]
    st.markdown("**상태 요약**")
    cards = [("신규 업체", f"{len(nc)}곳", True)] + [(k, str(stt.get(k, 0)), False) for k in ["완납", "진행", "미수", "장기미수"]]
    h = "<div style='display:grid;grid-template-columns:repeat(5,1fr);gap:10px;margin:4px 0 14px;'>"
    for lab, val, hi in cards:
        bg = NAVY if hi else "#EAF1FB"; lc = "#ccdcf5" if hi else "#3A5C8A"; vc = "#ffffff" if hi else NAVY
        h += f"<div style='background:{bg};border-radius:10px;padding:12px 14px;'><div style='font-size:12px;color:{lc};'>{lab}</div><div style='font-size:24px;font-weight:700;color:{vc};'>{val}</div></div>"
    st.markdown(h + "</div>", unsafe_allow_html=True)
    # 이번에 갱신·신규 처리된 거래처 (접기/펼치기)
    chg = [{"구분": s[8], "지역": s[0], "거래처": s[1], "상태": s[3],
            "계산서 추가": int(s[6]), "입금 추가": int(s[7])}
           for s in R["summary"] if (s[6] or s[7])]
    if chg:
        with st.expander(f"이번에 갱신·신규 처리된 거래처 {len(chg)}곳 자세히 보기 / 내려받기"):
            cdf = pd.DataFrame(chg).sort_values(["구분", "지역", "거래처"]).reset_index(drop=True)
            st.dataframe(cdf, use_container_width=True, hide_index=True)
            st.download_button("이 목록 CSV 다운로드", cdf.to_csv(index=False).encode("utf-8-sig"),
                               file_name="갱신_신규_거래처.csv", mime="text/csv", key="dl_chg")
    else:
        st.caption("이번에 갱신·신규 처리된 거래처가 없습니다. (동일 자료 재처리 등)")
    # 미배정
    if R["unassigned"]:
        st.warning(f"미배정 입금 {len(R['unassigned'])}건 (입금자명 매칭 실패 — 별칭표 보완 필요)")
        with st.expander(f"미배정 입금 {len(R['unassigned'])}건 자세히 보기 / 내려받기"):
            udf = pd.DataFrame(R["unassigned"], columns=["지역", "날짜", "금액", "입금자명", "유형"])
            udf = udf.sort_values(["지역", "입금자명", "날짜"]).reset_index(drop=True)
            st.dataframe(udf, use_container_width=True, hide_index=True)
            st.download_button("미배정 목록 CSV 다운로드", udf.to_csv(index=False).encode("utf-8-sig"),
                               file_name="미배정입금.csv", mime="text/csv", key="dl_un")
            if ADMIN:
                _render_alias_editor(udf)
    # 최종 결과 + 다운로드
    if R["ok"]:
        st.success("검증 통과 — 누적본을 갱신했고 직전본을 백업했습니다." + R["saved_note"])
        st.download_button("📥 갱신된 누적본 다운로드 (zip)", R["zip_bytes"],
                           file_name="누적자료_최신본.zip", mime="application/zip", type="primary", key="dl_final")
    else:
        st.error("검증 실패 — 누적본을 갱신하지 않았습니다. 직전본은 그대로 유지됩니다.")
        st.write(R["probs"][:15])
        st.download_button("⚠ 검증 전 결과 받아보기 (zip)", R["zip_bytes"],
                           file_name="누적자료_검증전.zip", mime="application/zip", key="dl_final")

# ===== 미수 현황 대시보드 페이지 =====
@st.cache_data(show_spinner="미수 현황 계산 중…")
def _dashboard_html(_fp, basis_iso, admin):
    import dashboard as _dash
    return _dash.render(DATA, datetime.date.fromisoformat(basis_iso), _duerules(), admin=admin)

@st.cache_data(show_spinner=False)
def _longoverdue_cached(_fp):
    return _longoverdue_list()

@st.cache_resource(show_spinner=False)
def _nyung_book(_fp):
    import nyung
    return nyung.load_book(_REPO_DIR)

def _data_fp():
    return (os.path.getmtime(DATA_ZIP) if os.path.exists(DATA_ZIP) else 0,
            len(glob.glob(os.path.join(DATA, "*", "*.xlsx"))))

def _tpl_mtime(name):
    """템플릿(HTML) 수정시각 — 캐시 키에 넣어, 템플릿만 바꿔도(재배포 시) 화면이 새로 렌더되게 함."""
    try:
        return os.path.getmtime(os.path.join(HERE, name))
    except Exception:
        return 0

@st.cache_data(show_spinner=False)
def _sales_summary_data(_fp):
    try:
        import sales as _sales
        bd = _sales.build_data(DATA, _REPO_DIR)
        if not bd: return {}
        years = bd.get("years") or {}; summ = bd.get("summary") or {}
        if not years: return {}
        cy = max(years.keys())
        sby = summ.get("매출 (계산서 기준)", {}); pby = summ.get("매출총이익", {}); mby = summ.get("마진율", {})
        months = [e for e in years.get(cy, []) if e.get("매출합")]
        lm = months[-1] if months else None
        buyby = summ.get("매입 (계산서 기준)", {})
        yrs = sorted(years.keys())[-6:]
        yearly = [{"y": y, "sales": int(sby.get(y, 0) or 0), "buy": int(buyby.get(y, 0) or 0)} for y in yrs]
        monthly = [{"m": e.get("월"), "sales": int(e.get("매출합", 0) or 0), "buy": int(e.get("매입합", 0) or 0)} for e in years.get(cy, [])]
        return {
            "sales_year": cy,
            "sales_year_total": int(sby.get(cy, 0) or 0),
            "sales_year_profit": int(pby.get(cy, 0) or 0),
            "sales_year_margin": round(float(mby.get(cy, 0) or 0) * 100, 1),
            "sales_latest_month": (lm.get("월") if lm else ""),
            "sales_latest_month_total": int(lm.get("매출합", 0)) if lm else 0,
            "sales_yearly": yearly,
            "sales_monthly": monthly,
        }
    except Exception:
        return {}

@st.cache_data(show_spinner=False)
def _misu_summary_data(_fp, basis_iso):
    import dashboard as _dash
    bd = datetime.date.fromisoformat(basis_iso)
    data = _dash.compute_data(DATA, bd, _duerules())
    unpaid = [o for o in data if o["status"] in ("미수", "장기미수")]
    longx = [o for o in data if o["status"] == "장기미수"]
    return {
        "basis": basis_iso,
        "updatedAt": datetime.datetime.now().isoformat(timespec="seconds"),
        "customer_count": len(data),
        "unpaid_count": len(unpaid),
        "longoverdue_count": len(longx),
        "total_unpaid": int(sum(o["amt"] for o in unpaid)),
        "seoul_unpaid": int(sum(o["amt"] for o in unpaid if o["reg"] == "서울")),
        "hwaseong_unpaid": int(sum(o["amt"] for o in unpaid if o["reg"] == "화성")),
        "longoverdue_amt": int(sum(o["amt"] for o in longx)),
    }

def page_dashboard():
    bd = basis_date() or datetime.date.today()
    if not os.path.isdir(DATA):
        header(); st.error("자료를 불러올 수 없습니다."); return
    try:
        _fp = _data_fp() + (_tpl_mtime("dashboard_template.html"),)
        _components.html(_dashboard_html(_fp, bd.isoformat(), ADMIN), height=760, scrolling=True)
        # 팝업의 '내용증명 작성' 버튼이 같은 출처(same-origin)로 클릭할 숨은 트리거(거래처별). 관리자·실무자 모두.
        with st.container(key="nytrig"):
            for _o in _longoverdue_cached(_fp):
                if st.button("NYO_" + _o["biz"], key="nyo_" + _o["biz"]):
                    st.session_state["_routed_biz"] = _o["biz"]
                    st.session_state["_ny_applied"] = None
                    st.switch_page(_nyung_page)
    except Exception as e:
        header(); st.error(f"대시보드 생성 오류: {e}")
    try:
        _sum = dict(_misu_summary_data(_fp, bd.isoformat()))
        try: _sum.update(_sales_summary_data(_fp))
        except Exception: pass
        _publish_misu_summary(_sum)
    except Exception:
        pass

# ===== 매출 현황 페이지 =====
@st.cache_data(show_spinner="매출 현황 불러오는 중…")
def _sales_html(_fp, basis_iso, admin):
    import sales as _sales
    return _sales.render(DATA, admin=admin, basis_iso=basis_iso)

def page_sales():
    if not os.path.isdir(DATA):
        header(title="매출 현황"); st.error("자료를 불러올 수 없습니다."); return
    if not os.path.exists(os.path.join(DATA, "_매출자료.json")):
        header(title="매출 현황")
        st.warning("매출 자료(_매출자료.json)가 없습니다. chjk-data의 data.zip에 매출 자료를 포함해 주세요.")
        return
    try:
        _bd = basis_date()
        _fp = (os.path.getmtime(DATA_ZIP) if os.path.exists(DATA_ZIP) else 0, _tpl_mtime("sales_template.html"))
        _components.html(_sales_html(_fp, _bd.isoformat() if _bd else "", ADMIN), height=900, scrolling=True)
    except Exception as e:
        header(title="매출 현황"); st.error(f"매출 현황 생성 오류: {e}")

# ===== 내용증명 페이지 =====
def _longoverdue_list():
    """현재 장기미수 거래처(실시간 계산). 상태가 바뀌면 자동 반영."""
    import dashboard as _dash
    bd = basis_date() or datetime.date.today()
    data = _dash.compute_data(DATA, bd, _duerules())
    lt = [o for o in data if o.get("status") == "장기미수"]
    lt.sort(key=lambda o: (o["reg"], -o["amt"]))
    return lt

def _fontconfig_file():
    """맑은 고딕/Malgun Gothic → Noto Sans CJK KR 로 강제 매핑(서버 PDF가 워드와 동일하게 보이도록)."""
    conf = os.path.join(tempfile.gettempdir(), "chjk_fonts.conf")
    try:
        if not os.path.exists(conf):
            with open(conf, "w", encoding="utf-8") as f:
                f.write('<?xml version="1.0"?>\n<!DOCTYPE fontconfig SYSTEM "fonts.dtd">\n<fontconfig>\n'
                        '  <include ignore_missing="yes">/etc/fonts/fonts.conf</include>\n'
                        '  <match target="pattern"><test name="family"><string>맑은 고딕</string></test>'
                        '<edit name="family" mode="assign" binding="strong"><string>Noto Sans CJK KR</string></edit></match>\n'
                        '  <match target="pattern"><test name="family"><string>Malgun Gothic</string></test>'
                        '<edit name="family" mode="assign" binding="strong"><string>Noto Sans CJK KR</string></edit></match>\n'
                        '</fontconfig>\n')
        return conf
    except Exception:
        return None

def _docx_to_pdf(docx_path, out_dir):
    """LibreOffice headless 로 docx→pdf. 폰트 별칭(맑은고딕→Noto)으로 워드와 동일하게 렌더. 실패 시 None."""
    import subprocess
    base = os.path.splitext(os.path.basename(docx_path))[0]
    env = dict(os.environ)
    conf = _fontconfig_file()
    if conf:
        env["FONTCONFIG_FILE"] = conf
    lohome = os.path.join(tempfile.gettempdir(), "lohome")
    try:
        os.makedirs(lohome, exist_ok=True); env.setdefault("HOME", lohome)
    except Exception:
        pass
    for soffice in ("libreoffice", "soffice"):
        try:
            subprocess.run([soffice, "--headless", "--convert-to", "pdf", "--outdir", out_dir, docx_path],
                           check=True, timeout=120, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, env=env)
            pdf = os.path.join(out_dir, base + ".pdf")
            if os.path.exists(pdf):
                return pdf
        except Exception:
            continue
    return None

def _pdf_pages(path):
    try:
        from pypdf import PdfReader
        return len(PdfReader(path).pages)
    except Exception:
        return 1

def _reformat_amt(_k):
    """미수액 입력값을 천단위 콤마로 자동 정리(on_change 콜백 — 위젯 키 수정 허용 구간)."""
    d = re.sub(r"[^0-9]", "", str(st.session_state.get(_k, "")))
    st.session_state[_k] = f"{int(d):,}" if d else ""

def _make_pdf_fit(data, pdf_src, workdir):
    """PDF는 노토 폰트로 생성. 항목간격을 키워 하단까지 채우되, 페이지수를 보며 1페이지를 유지하도록 자동 조절.
    서버 렌더링이 sandbox보다 촘촘해도 알아서 더 채운다. 성공값은 세션에 캐시해 다음부터 빠르게."""
    import nyung
    cached = st.session_state.get("_pdf_sp")
    cands = ([cached] if cached else []) + [9, 7, 6, 5, 4, 3]
    seen = set(); order = [c for c in cands if not (c in seen or seen.add(c))]
    last = None
    for sp in order:
        # 본문 번호 사이(1-2,2-3,3-4,4-5,6-7)와 구간 사이(수신자↔발신자, 발신자↔촉구의 건, 입금표 위·아래, 실선↔날짜, 회사명↔명판)를 균일하게 띄워 고르게 채움
        nyung.make_docx(data, _REPO_DIR, pdf_src, font="Noto Sans CJK KR", date_before=2, item_after=sp, spread=sp)
        cand = _docx_to_pdf(pdf_src, workdir)
        if not cand:
            return None
        last = cand
        if _pdf_pages(cand) == 1:
            st.session_state["_pdf_sp"] = sp
            return cand
    return last

def page_nyung():
    import nyung
    header(title="내용증명")
    if st.button("← 미수 현황으로 돌아가기"):
        st.switch_page(_dashboard_page)
    if not os.path.isdir(DATA):
        st.error("자료를 불러올 수 없습니다."); return
    miss = [f for f in ("stamp_guro.png", "stamp_hwaseong.png", "인감_보정_투명.png")
            if not os.path.exists(os.path.join(_REPO_DIR, f))]
    if miss:
        st.warning("명판/인감 파일이 자료 저장소(chjk-data)에 없습니다: " + ", ".join(miss)
                   + "  · PDF/명판이 비정상일 수 있습니다.")
    with st.spinner("불러오는 중…"):
        _fp = _data_fp()
        lt = _longoverdue_cached(_fp)
        book = _nyung_book(_fp)
    if not lt:
        st.info("현재 장기미수 거래처가 없습니다."); return
    qbiz = st.session_state.get("_routed_biz") or st.query_params.get("biz", "")
    labels = [f"[{o['reg']}] {o['name']} · {o['amt']:,}원" for o in lt]
    idx = next((i for i, o in enumerate(lt) if o["biz"] == qbiz), 0)
    # 팝업에서 새 거래처가 넘어오면 셀렉트박스를 그 거래처로 강제(이후엔 사용자 선택 유지)
    if qbiz and st.session_state.get("_ny_applied") != qbiz:
        st.session_state["_ny_applied"] = qbiz
        st.session_state["ny_sel"] = idx
    sel = st.selectbox("장기미수 거래처 선택", range(len(lt)), index=idx,
                       format_func=lambda i: labels[i], key="ny_sel")
    o = lt[sel]; biz = o["biz"]
    m = nyung.match_company(book, o["name"], biz) or {}
    st.caption("자동완성된 내용을 그대로 쓰거나 수정 후 생성하세요. 거래처를 바꾸면 자동으로 다시 채워집니다.")
    c1, c2 = st.columns([3, 2])
    corp = c1.text_input("거래처명 (정식 상호)", value=(m.get("name") or o["name"]), key=f"corp_{biz}")
    _amk = f"amt_{biz}"
    if _amk not in st.session_state:
        st.session_state[_amk] = f"{int(round(o['amt'])):,}"
    amt_raw = c2.text_input("미수액 (원)", key=_amk, on_change=_reformat_amt, args=(_amk,),
                            help="천 단위 쉼표가 자동으로 표시됩니다.")
    amt = int(re.sub(r"[^0-9]", "", amt_raw or "") or 0)
    c3, c4 = st.columns(2)
    rep = c3.text_input("대표자명 (수신 담당자)", value=m.get("rep", ""), key=f"rep_{biz}")
    tel = c4.text_input("수신 연락처", value=m.get("tel", ""), key=f"tel_{biz}")
    addr = st.text_input("수신 주소", value=m.get("addr", ""), key=f"addr_{biz}")
    reg = st.radio("관할 (명판·발신자·입금계좌가 바뀝니다)", ["서울", "화성"],
                   index=(0 if o["reg"] == "서울" else 1), horizontal=True, key=f"reg_{biz}")
    if not m:
        st.info("이 거래처는 명단에 없어 대표자·연락처·주소가 비어 있습니다. 직접 입력하세요.")
    st.caption(f"미수액 미리보기: {amt:,}원 · 하단 날짜는 다운로드 당일이 자동 입력됩니다.")
    if st.button("📄 내용증명 생성 (워드 + PDF)", type="primary"):
        if not corp.strip():
            st.error("거래처명을 입력하세요.")
        else:
            data = dict(거래처명=corp.strip(), 담당자=rep.strip(), 수신전화=tel.strip(),
                        수신주소=addr.strip(), 미수액=amt, 관할=reg)
            workdir = tempfile.mkdtemp(prefix="nyung_")
            safe = re.sub(r"[^가-힣A-Za-z0-9]", "_", corp.strip()) or "내용증명"
            docx_path = os.path.join(workdir, f"내용증명_{safe}.docx")
            try:
                with st.spinner("내용증명 생성 중… (PDF 하단 채움 최적화로 몇 초 더 걸릴 수 있어요)"):
                    nyung.make_docx(data, _REPO_DIR, docx_path)  # 워드: 맑은 고딕(이전 형식 그대로 유지)
                    wb = open(docx_path, "rb").read()
                    pdf_src = os.path.join(workdir, "pdf_src.docx")  # PDF: 별도 규칙(노토 + 하단까지 자동 채움)
                    pdf = _make_pdf_fit(data, pdf_src, workdir)
                    pb = open(pdf, "rb").read() if pdf else None
                st.session_state["ny_out"] = {"biz": biz, "word": wb, "pdf": pb, "name": f"내용증명_{safe}"}
            except Exception as e:
                st.session_state.pop("ny_out", None)
                st.error(f"문서 생성 오류: {e}")
    out = st.session_state.get("ny_out")
    if out and out.get("biz") == biz:
        st.success("생성 완료. 아래에서 내려받으세요.")
        d1, d2 = st.columns(2)
        d1.download_button("📥 워드(.docx) 받기", out["word"], file_name=out["name"] + ".docx",
                           mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document", key="dl_word")
        if out.get("pdf"):
            d2.download_button("📥 PDF 받기", out["pdf"], file_name=out["name"] + ".pdf",
                               mime="application/pdf", key="dl_pdf")
        else:
            d2.warning("PDF 변환 실패 — 서버에 libreoffice-writer 가 필요합니다. 워드는 정상입니다.")

# ===== 미배정 입금 → 별칭 등록 =====
def _alias_cust_options():
    """지역별 거래처 목록. {지역: {표시라벨: 거래처명}}"""
    opts = {"서울": {}, "화성": {}}
    for loc, f, n in all_company_files():
        nm = re.sub(r"^[^)]*\)\s*", "", n).rsplit(" (", 1)[0]
        m = re.search(r"\((\d{3}-\d{2}-\d{5})\)", n)
        opts.setdefault(loc, {})[f"{nm} ({m.group(1)})" if m else nm] = nm
    return opts

def _render_alias_editor(udf):
    """미배정 입금자명마다 거래처를 지정 → _거래처별칭표.xlsx 에 저장 → chjk-data 반영."""
    st.markdown("---")
    st.markdown("**이 입금자명을 앞으로 자동 매칭시키기**")
    st.caption("입금자명마다 거래처를 지정하면 별칭표에 저장되어, 다음 업로드부터 자동 매칭됩니다. "
               "거래처 입금이 아니면(세금·이자·수수료 등) '제외'를 고르세요.")
    opts = _alias_cust_options()
    uniq = udf[["지역", "입금자명"]].drop_duplicates().reset_index(drop=True)
    picks = {}
    for i, row in uniq.iterrows():
        loc = str(row["지역"]); nm = str(row["입금자명"])
        labels = ["(지정 안 함)", "❌ 제외 (거래처 입금 아님)"] + sorted(opts.get(loc, {}).keys())
        c1, c2 = st.columns([1, 2])
        c1.markdown(f"<div style='padding-top:7px'><b>{loc}</b> · {nm}</div>", unsafe_allow_html=True)
        sel = c2.selectbox("거래처", labels, index=0, key=f"alias_{loc}_{i}", label_visibility="collapsed")
        if sel and sel != "(지정 안 함)":
            picks[(loc, nm)] = "제외" if sel.startswith("❌") else opts[loc][sel]
    if st.button(f"💾 별칭표에 저장 ({len(picks)}건)", type="primary", disabled=not picks, key="save_alias"):
        try:
            rows = [(loc, nm, tg) for (loc, nm), tg in picks.items()]
            n = store.save_alias_rows(DATA, rows, DATA_ZIP)
            saved, msg = store.git_commit_push(_REPO_DIR, GIT_TOKEN, DATA_REPO, "add deposit aliases")
            st.success(f"{n}건을 별칭표에 저장했습니다." + ("" if saved else f" (저장소 반영 실패: {msg})"))
            st.info("이번 미배정 건까지 반영하려면 같은 은행 파일을 다시 올려 처리하세요. "
                    "이미 반영된 입금은 (거래일·금액) 기준으로 중복되지 않습니다.")
        except Exception as e:
            st.error(f"별칭 저장 실패: {e}")

# ===== 자료 처리 페이지 =====
def page_process():
    header(title="자료 처리")
    if ADMIN:
        st.markdown(f"<div style='margin:-6px 0 8px;'><span style='background:{NAVY};color:#fff;font-size:12px;font-weight:600;padding:3px 12px;border-radius:6px;'>🔑 관리자 계정</span></div>", unsafe_allow_html=True)
        st.caption("매출·매입 세금계산서, 어음수취내역, 은행거래내역을 올리면 자동으로 종류를 판별하고 기존 자료에 신규만 추가합니다.")
    if DATA_REPO and GIT_TOKEN and not _data_repo_dir():
        st.error("⚠ 자료 저장소(chjk-data) 연결 실패 — Secrets의 github_token 권한(chjk-data Contents: Read and write)과 github_data_repo 값을 확인하세요.")

    with st.expander("🔎 거래처 검색 · 개별 다운로드", expanded=False):
        files = all_company_files()
        _label_map = {}
        for loc, f, n in files:
            _label_map[_cust_label(loc, n)] = (f, n)
        _opts = list(_label_map.keys())
        sel = st.selectbox(f"거래처명 일부만 입력하면 자동으로 후보가 나옵니다 (전체 {len(files)}곳) · 엔터(또는 클릭) 후 바로 입력하면 새로 검색됩니다",
                           _opts, index=None, placeholder="거래처명 입력 (예: 농협)", key="cust_sel")
        if sel and sel in _label_map:
            f, n = _label_map[sel]
            with open(f, "rb") as fh:
                st.download_button(f"📥 {sel} 다운로드", fh.read(), file_name=n, key="dl_sel",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    with st.expander("📥 현재 누적자료 전체 다운로드 (zip)", expanded=False):
        if os.path.isdir(DATA):
            st.download_button("전체 누적본 다운로드 (서울·화성 거래처)", zip_bytes(DATA, ["서울", "화성"]), file_name="누적자료_현재본.zip", mime="application/zip", key="dl_all")

    with st.expander("🛟 비상 복구 (직전본)", expanded=bool(st.session_state.get("confirm_restore"))):
        if os.path.exists(BACKUP_ZIP):
            st.caption("최근 갱신 직전 상태가 백업되어 있습니다. 문제가 생기면 한 번에 되돌릴 수 있어요.")
            bc1, bc2 = st.columns(2)
            bc1.download_button("직전본 다운로드 (서울·화성)", zip_backup_customers(), file_name="누적자료_직전본.zip", mime="application/zip", key="dl_bak")
            if ADMIN and bc2.button("⏪ 직전본으로 복구"):
                st.session_state["confirm_restore"] = True
            if ADMIN and st.session_state.get("confirm_restore"):
                st.warning("직전본으로 되돌리면 현재 누적본이 직전 상태로 바뀝니다. 정말 실행하시겠습니까?")
                rc1, rc2 = st.columns(2)
                if rc1.button("예, 복구 실행", type="primary", key="do_restore"):
                    st.session_state.pop("confirm_restore", None)
                    if store.restore_previous(DATA, DATA_ZIP, BACKUP_ZIP):
                        if GIT_TOKEN: store.git_commit_push(_REPO_DIR, GIT_TOKEN, DATA_REPO, "restore previous baseline")
                        st.session_state.pop("result", None)
                        st.success("직전본으로 복구했습니다."); st.rerun()
                if rc2.button("취소", key="cancel_restore"):
                    st.session_state.pop("confirm_restore", None); st.rerun()
        else:
            st.caption("아직 직전본 백업이 없습니다. (첫 갱신 후 생성됩니다)")

    if ADMIN:
        with st.expander("🔄 전체 다시 계산 (설정 변경 반영)"):
            st.caption("결제조건·별칭 등 설정만 바꿨을 때, 업로드 없이 전체 거래처의 상태·파일명을 현재 기준일로 다시 계산합니다.")
            if st.button("전체 다시 계산 실행", key="recalc_all"):
                old = {}
                for loc in ["서울", "화성"]:
                    for f in glob.glob(os.path.join(DATA, loc, "*.xlsx")):
                        nm = os.path.basename(f)
                        if nm.startswith("_"): continue
                        mb = re.search(r"(\d{3}-\d{2}-\d{5})", nm)
                        if mb:
                            old[mb.group(1)] = ("완납" if "완납" in nm else "장기미수" if "장기미수" in nm else "미수" if "미수" in nm else "진행")
                work = tempfile.mkdtemp(); out_dir = os.path.join(work, "out"); os.makedirs(out_dir, exist_ok=True)
                with st.spinner("전체 거래처 다시 계산 중… (수십 초)"):
                    try:
                        res = pipeline.process({"서울": [], "화성": []}, DATA, out_dir, ref_date=basis_date())
                        ok, probs = pipeline.verify(out_dir, DATA)
                    except Exception as e:
                        ok = False; probs = [str(e)]
                if ok:
                    changes = [(s[0], s[1], old.get(s[2], "?"), s[3]) for s in res["summary"] if old.get(s[2], s[3]) != s[3]]
                    store.apply_update(out_dir, DATA, DATA_ZIP, BACKUP_ZIP)
                    saved, msg = store.git_commit_push(_REPO_DIR, GIT_TOKEN, DATA_REPO, "recompute all statuses")
                    st.cache_data.clear()
                    note = "  (GitHub 영구저장 완료)" if saved else f"  (GitHub 미저장: {msg})"
                    if changes:
                        st.success(f"전체 다시 계산 완료 — {len(changes)}곳 상태가 바뀌었습니다." + note)
                        cdf = pd.DataFrame(changes, columns=["지역", "거래처", "이전", "변경"]).sort_values(["지역", "거래처"]).reset_index(drop=True)
                        st.dataframe(cdf, use_container_width=True, hide_index=True)
                    else:
                        st.success("전체 다시 계산 완료 — 바뀐 상태가 없습니다." + note)
                else:
                    st.error("재계산 검증 실패 — 반영하지 않았습니다."); st.write(probs[:10])

        _md = st.session_state.pop("manual_done", None)
        with st.expander("✏️ 거래처 파일 직접 수정 후 업로드 (수동 교체)", expanded=bool(_md)):
            if _md: st.success(_md)
            st.caption("거래처 파일을 받아 직접 고친 뒤 여기 올리면 그 거래처를 교체합니다. 사업자번호로 자동 인식하며, 교체 전 직전본이 백업됩니다.")
            m_ups = st.file_uploader("수정한 거래처 파일(.xlsx) 업로드", accept_multiple_files=True, type=["xlsx"], key="manual_edit")
            if m_ups:
                parsed = []
                for uf in m_ups:
                    nm = uf.name; raw = bytes(uf.getbuffer())
                    mb = re.search(r"\((\d{3}-\d{2}-\d{5})\)", nm)
                    ml = re.search(r"(서울|화성)\)", nm)
                    loc = ml.group(1) if ml else ("화성" if "화성" in nm else "서울")
                    try: load_workbook(io.BytesIO(raw)); ok = True
                    except Exception: ok = False
                    parsed.append((loc, nm, (mb.group(1) if mb else None), ok, raw))
                for loc, nm, biz, ok, _ in parsed:
                    st.write(f"- [{loc}] {nm} → " + ("✅ 교체 준비됨" if (ok and biz) else "⚠️ 사업자번호/형식 확인 필요(건너뜀)"))
                good = [(loc, nm, b) for loc, nm, biz, ok, b in parsed if ok and biz]
                if good and st.button(f"{len(good)}개 거래처 교체 (직전본 백업 후)", type="primary", key="manual_apply"):
                    store.replace_customer_files(DATA, good, DATA_ZIP, BACKUP_ZIP)
                    saved, msg = store.git_commit_push(_REPO_DIR, GIT_TOKEN, DATA_REPO, "manual edit replace")
                    st.session_state.pop("result", None)
                    st.session_state["manual_done"] = f"✅ {len(good)}개 거래처 교체 완료." + ("  (GitHub 영구저장 완료)" if saved else f"  (GitHub 미저장: {msg})")
                    st.toast(f"{len(good)}개 거래처 교체 완료 ✅")
                    st.rerun()

        st.subheader("자료 업로드")
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("**서울** (신한·기업·하나·농협·우리 등)")
            up_seoul = st.file_uploader("서울 자료", accept_multiple_files=True, key="seoul",
                                        type=["xls", "xlsx"], label_visibility="collapsed")
        with col2:
            st.markdown("**화성** (신한)")
            up_hwa = st.file_uploader("화성 자료", accept_multiple_files=True, key="hwa",
                                      type=["xls", "xlsx"], label_visibility="collapsed")

        if st.button("🚀 처리 시작", type="primary", disabled=not (up_seoul or up_hwa)):
            work = tempfile.mkdtemp(); up_dir = os.path.join(work, "up"); out_dir = os.path.join(work, "out")
            os.makedirs(up_dir, exist_ok=True); os.makedirs(out_dir, exist_ok=True)
            uploads = {"서울": [], "화성": []}
            for loc, ulist in [("서울", up_seoul or []), ("화성", up_hwa or [])]:
                for f in ulist:
                    p = os.path.join(up_dir, f"{loc}_{f.name}"); open(p, "wb").write(f.getbuffer()); uploads[loc].append(p)
            log_area = st.empty(); logs = []
            def prog(m): logs.append(m); log_area.code("\n".join(logs[-12:]))
            with st.spinner("처리 중… (파일이 많으면 수십 초 걸릴 수 있어요)"):
                try:
                    res = pipeline.process(uploads, DATA if os.path.isdir(DATA) else up_dir, out_dir, progress=prog, ref_date=basis_date())
                except Exception as e:
                    st.error(f"처리 오류: {e}"); st.stop()
                ok, probs = pipeline.verify(out_dir, DATA)
                if ok:
                    store.apply_update(out_dir, DATA, DATA_ZIP, BACKUP_ZIP)
                    open(BASIS, "w", encoding="utf-8").write(res.get("basis") or datetime.date.today().isoformat())
                    saved, msg = store.git_commit_push(_REPO_DIR, GIT_TOKEN, DATA_REPO, "update baseline via app")
                    note = "  (GitHub 영구저장 완료)" if saved else f"  (GitHub 미저장: {msg})"
                    zb = zip_bytes(DATA)
                else:
                    note = ""; zb = zip_bytes(out_dir)
            st.session_state["result"] = {
                "detected": res["detected"], "status": res["status"], "new_companies": res.get("new_companies", []),
                "summary": res["summary"], "unassigned": res["unassigned"],
                "ok": ok, "probs": probs, "zip_bytes": zb, "saved_note": note,
            }
            log_area.empty()

        if st.session_state.get("result"):
            render_result(st.session_state["result"])

if not check_pw(): st.stop()
ADMIN = st.session_state.get("role") == "admin"
if st.session_state.get("uid"):
    import urllib.parse as _up2
    _lo_url = _FB_DB_URL + "/" + _FB_PATH + "_logout/" + _up2.quote(st.session_state.get("uid",""), safe="") + ".json"
    _ck_js = ("<script>try{window.parent.localStorage.setItem('misu_last_user'," + json.dumps(st.session_state.get("uid","")) + ");"
              + "var d=window.parent.document;"
              + "d.cookie='misu_rt='+encodeURIComponent(" + json.dumps(st.session_state.get("_fb_ref","")) + ")+';path=/;max-age=2592000;SameSite=Lax';"
              + "d.cookie='misu_uid='+encodeURIComponent(" + json.dumps(st.session_state.get("uid","")) + ")+';path=/;max-age=2592000;SameSite=Lax';"
              + "d.cookie='misu_at=" + str(int(st.session_state.get("_login_at_ms") or 0)) + ";path=/;max-age=2592000;SameSite=Lax';"
              + "try{var L=window.parent.localStorage;L.setItem('misu_rt'," + json.dumps(st.session_state.get("_fb_ref","")) + ");L.setItem('misu_uid'," + json.dumps(st.session_state.get("uid","")) + ");L.setItem('misu_at','" + str(int(st.session_state.get("_login_at_ms") or 0)) + "');window.parent.sessionStorage.removeItem('misu_auto_tried');}catch(_l){}"
              + "}catch(e){}"
              # 실시간 로그아웃 감지: 5초마다 로그아웃 기록(수 바이트)만 확인
              + "var _at=" + str(int(st.session_state.get("_login_at_ms") or 0)) + ";"
              + "var _tok=" + json.dumps(st.session_state.get("_fb_tok","")) + ";"
              + "var _lt=setInterval(function(){"
              + "fetch(" + json.dumps(_lo_url) + "+'?auth='+_tok).then(function(r){return r.ok?r.json():null;}).then(function(v){"
              + "if(typeof v==='number'){clearInterval(_lt);"
              + "try{var d=window.parent.document;d.cookie='misu_rt=;path=/;max-age=0';d.cookie='misu_uid=;path=/;max-age=0';d.cookie='misu_at=;path=/;max-age=0';var L=window.parent.localStorage;L.removeItem('misu_rt');L.removeItem('misu_uid');L.removeItem('misu_at');}catch(e){}"
              + "try{var _pd2=window.parent.document;var _sc2=_pd2.createElement('script');_sc2.textContent='location.reload();';(_pd2.head||_pd2.body).appendChild(_sc2);}catch(_e2){try{window.parent.location.reload();}catch(_e3){}}}"
              + "}).catch(function(){});"
              + "},5000);"
              + "</script>")
    _components.html(_ck_js, height=0)
# 페이지 객체(전역) — 팝업 트리거와 '돌아가기'에서 st.switch_page 로 이동. 미수현황 탭은 항상 홈으로 복귀.
_dashboard_page = st.Page(page_dashboard, title="미수 현황", icon="📊", default=True)
_sales_page = st.Page(page_sales, title="매출 현황", icon="📈")
_process_page = st.Page(page_process, title="자료 처리", icon="🗂")
_nyung_page = st.Page(page_nyung, title="내용증명", icon="📄", url_path="page_nyung")
_pages = [_dashboard_page, _sales_page, _nyung_page]   # 내용증명은 관리자·실무자 모두(탭은 CSS로 숨김, 팝업으로 진입)
_pages.append(_process_page)
_pg = st.navigation(_pages, position="top")
_pg.run()
