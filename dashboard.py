# -*- coding: utf-8 -*-
"""미수 현황 대시보드: 거래처 파일에서 실시간 계산해 HTML 렌더.
상태/미수액은 기준일(=_기준일.txt, 자료 최근거래일) 기준. 비고 '수취 불가/부도' 입금은 엔진이 자동 제외."""
import os, glob, re, json, datetime
from openpyxl import load_workbook
import engine_core as E

def _fmt(x):
    if x >= 1e8: return f"{x/1e8:.1f}억"
    if x >= 1e4: return f"{round(x/1e4):,}만"
    return f"{x:,.0f}"

def _cust(path, reg, basis, duerules):
    b = os.path.basename(path)
    m = re.search(r'(\d{3}-\d{2}-\d{5})', b)
    if not m: return None
    bn = m.group(1)
    nm = re.sub(r'^[^)]*\)\s*', '', b).rsplit(' (', 1)[0]
    wb = load_workbook(path); ws = wb.active; C = E.cols(ws)
    supmap = {}; deps = []
    for r in range(3, ws.max_row + 1):
        a = E.pdate(ws.cell(r, C['작성']).value); sv = ws.cell(r, C['공급']).value
        if a and isinstance(sv, (int, float)): supmap[a.isoformat()] = supmap.get(a.isoformat(), 0) + sv
        bg = ws.cell(r, C['비고']).value
        if bg and E.NORECV.search(str(bg)): continue   # 부도/수취불가 입금은 이력에서도 제외
        dd = E.pdate(ws.cell(r, C['입금일']).value); dm = ws.cell(r, C['입금']).value
        if dd and isinstance(dm, (int, float)) and dm > 0: deps.append((dd.isoformat(), float(dm)))
    worst = '완납'; total = 0
    for a, bal in E.unpaid_after_credit(ws, C):
        total += bal
        st = E.status(a, today=basis, rule=E.resolve_rule(duerules.get(bn, '익월말'), supmap.get(a.isoformat(), 0)))
        worst = '장기미수' if st == '장기미수' else ('미수' if st == '미수' and worst != '장기미수' else ('진행' if st == '기한내' and worst == '완납' else worst))
    wb.close(); deps.sort()
    rule = duerules.get(bn, '익월말'); ruledisp = '조건부' if '?' in rule else rule
    last = deps[-1][0] if deps else None
    since = (basis - datetime.date.fromisoformat(last)).days if last else None
    return {'name': nm, 'reg': reg, 'biz': bn, 'status': worst, 'amt': round(total),
            'last_dep': last, 'since': since, 'rule': rule, 'ruledisp': ruledisp,
            'deps': [{'d': d, 'a': a} for d, a in deps[-8:]]}

def compute_data(data_dir, basis, duerules):
    out = []
    for reg in ['서울', '화성']:
        for path in sorted(glob.glob(os.path.join(data_dir, reg, '*.xlsx'))):
            b = os.path.basename(path)
            if b.startswith('_') or b.startswith('~$'): continue
            try:
                o = _cust(path, reg, basis, duerules)
                if o: out.append(o)
            except Exception:
                continue
    return out

def _bars(unpaid, reg, color):
    rows = sorted([o for o in unpaid if o['reg'] == reg], key=lambda o: -o['amt'])[:5]
    mx = max([o['amt'] for o in rows] + [1])
    head = '<div class=cbhead><span>거래처</span><span>미수액 · 최근입금일</span></div>'
    bars = "".join(
        "<div class=cbrow><div class=cbname title=\"" + o['name'] + "\">" + o['name'] +
        "</div><div class=cbbar><div class=cbfill style='width:" + str(max(7, round(o['amt'] / mx * 100))) +
        "%;background:" + color + "'></div></div><div class=cbval>" + _fmt(o['amt']) +
        "<div class=cbdate>" + (o['last_dep'] or '입금없음') + "</div></div></div>"
        for o in rows)
    return rows, head + bars

def render(data_dir, basis, duerules, template_path=None, admin=False):
    """basis: datetime.date (기준일=_기준일.txt). 반환: 대시보드 HTML 문자열."""
    data = compute_data(data_dir, basis, duerules)
    unpaid = [o for o in data if o['status'] in ('미수', '장기미수')]
    sr, sb = _bars(unpaid, '서울', '#1B3A6B')
    hr, hb = _bars(unpaid, '화성', '#3A6BB0')
    st_tot = sum(o['amt'] for o in unpaid if o['reg'] == '서울')
    hw_tot = sum(o['amt'] for o in unpaid if o['reg'] == '화성')
    charts = ("<div class=charts>\n"
              " <div class=chcard onclick=\"setRegion('서울')\"><h3>서울 대표 미수처 <span>· 총 " +
              _fmt(st_tot) + "원 (상위 " + str(len(sr)) + ")</span></h3>" + sb + "</div>\n"
              " <div class=chcard onclick=\"setRegion('화성')\"><h3>화성 대표 미수처 <span>· 총 " +
              _fmt(hw_tot) + "원 (상위 " + str(len(hr)) + ")</span></h3>" + hb + "</div>\n</div>\n")
    tp = template_path or os.path.join(os.path.dirname(os.path.abspath(__file__)), 'dashboard_template.html')
    tpl = open(tp, encoding='utf-8').read()
    bk = f"{basis.year}년 {basis.month}월 {basis.day}일"
    badge = ("<div style='margin:0 0 12px'><span style='background:#1B3A6B;color:#fff;font-size:12px;"
             "font-weight:600;padding:3px 12px;border-radius:6px'>🔑 관리자 계정</span></div>") if admin else ""
    return (tpl.replace('@@DATA@@', json.dumps(data, ensure_ascii=False))
               .replace('@@CHARTS@@', charts)
               .replace('@@BASIS@@', bk)
               .replace('@@ADMINBADGE@@', badge))
