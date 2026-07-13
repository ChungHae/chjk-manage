# -*- coding: utf-8 -*-
"""누적본을 zip 1개(data.zip)로 영구저장. 실행 시 작업폴더로 풀고, 갱신 시 _직전본.zip 백업 후 재압축, 선택적 GitHub 커밋."""
import os, shutil, glob, subprocess, zipfile

def _zip_dir(src_dir, zip_path):
    tmp = zip_path + ".tmp"
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as z:
        for root, _, files in os.walk(src_dir):
            for fn in files:
                fp = os.path.join(root, fn)
                z.write(fp, os.path.relpath(fp, src_dir))
    os.replace(tmp, zip_path)

def _unzip(zip_path, dst_dir):
    os.makedirs(dst_dir, exist_ok=True)
    if os.path.exists(zip_path):
        with zipfile.ZipFile(zip_path) as z: z.extractall(dst_dir)

def ensure_data(data_zip, work_dir):
    """실행 시 data.zip을 work_dir로 전개(이미 있으면 그대로)."""
    if (not os.path.isdir(work_dir)) or (not os.listdir(work_dir)):
        _unzip(data_zip, work_dir)
    return work_dir

def _replace(work_dir, src_dir):
    for sub in ("서울", "화성"):
        dst = os.path.join(work_dir, sub); src = os.path.join(src_dir, sub)
        if not os.path.isdir(src): continue
        if os.path.isdir(dst): shutil.rmtree(dst)
        shutil.copytree(src, dst)
    for t in glob.glob(os.path.join(src_dir, "_*.xlsx")): shutil.copy(t, work_dir)
    for t in glob.glob(os.path.join(src_dir, "_*.json")): shutil.copy(t, work_dir)

def apply_update(out_dir, work_dir, data_zip, backup_zip):
    """갱신 전 현재본을 _직전본.zip으로 백업 → work_dir 교체 → data.zip 재생성."""
    _zip_dir(work_dir, backup_zip)        # 직전본 = 갱신 전 상태
    _replace(work_dir, out_dir)
    _zip_dir(work_dir, data_zip)

def restore_previous(work_dir, data_zip, backup_zip):
    """_직전본.zip으로 한 번에 복구."""
    if not os.path.exists(backup_zip): return False
    shutil.rmtree(work_dir, ignore_errors=True)
    _unzip(backup_zip, work_dir)
    _zip_dir(work_dir, data_zip)
    return True

def git_commit_push(repo_dir, token, slug, msg):
    """GitHub 저장소에 되돌려 커밋·푸시(영구저장 + 모든 버전 git 이력 백업). token/slug 있을 때만."""
    if not token or not slug: return (False, "GitHub 미설정(다운로드만 가능)")
    try:
        subprocess.run(["git", "-C", repo_dir, "add", "-A"], check=True, capture_output=True)
        subprocess.run(["git", "-C", repo_dir, "-c", "user.email=app@chjk", "-c", "user.name=misu-app",
                        "commit", "-m", msg], check=True, capture_output=True)
        url = f"https://x-access-token:{token}@github.com/{slug}.git"
        subprocess.run(["git", "-C", repo_dir, "push", url, "HEAD"], check=True, capture_output=True, timeout=120)
        return (True, "저장 완료")
    except subprocess.CalledProcessError as e:
        return (False, (e.stderr or b"").decode("utf-8", "ignore")[:300])

def replace_customer_files(work_dir, items, data_zip, backup_zip):
    """관리자가 직접 수정한 거래처 파일로 교체. items=[(loc, filename, bytes)]. 교체 전 직전본 백업 후 data.zip 재생성.
    같은 사업자번호의 기존 파일은 삭제(상태 prefix가 달라져도 중복 안 생기게)."""
    import re as _re
    _zip_dir(work_dir, backup_zip)              # 교체 전 백업
    n = 0
    for loc, fn, content in items:
        d = os.path.join(work_dir, loc); os.makedirs(d, exist_ok=True)
        m = _re.search(r"\((\d{3}-\d{2}-\d{5})\)", fn)
        if m:
            for old in glob.glob(os.path.join(d, f"*({m.group(1)}).xlsx")):
                try: os.remove(old)
                except OSError: pass
        with open(os.path.join(d, fn), "wb") as f: f.write(content)
        n += 1
    _zip_dir(work_dir, data_zip)
    return n

def clone_data_repo(data_repo, token, dest):
    """비공개 자료 저장소(chjk-data)를 토큰으로 clone. 성공 시 경로 반환, 미설정/실패 시 None."""
    if not (data_repo and token): return None
    url = f"https://x-access-token:{token}@github.com/{data_repo}.git"
    try:
        shutil.rmtree(dest, ignore_errors=True)
        subprocess.run(["git", "clone", "--depth", "1", url, dest], check=True, capture_output=True, timeout=120)
        return dest
    except Exception:
        return None

def save_alias_rows(work_dir, rows, data_zip):
    """미배정 입금자명 → 거래처 매핑을 _거래처별칭표.xlsx에 추가/갱신한 뒤 data.zip 재생성.
    rows = [(지역, 입금자명, 대상거래처 or '제외')]
    같은 (지역, 정규화 입금자명)이 이미 있으면 값만 갱신한다.
    거래처 자료를 바꾸지 않으므로 _직전본.zip(복구지점)은 건드리지 않는다. 이력은 git 커밋으로 남는다."""
    if not rows: return 0
    from openpyxl import Workbook, load_workbook
    import engine_core as E
    p = os.path.join(work_dir, "_거래처별칭표.xlsx")
    if os.path.exists(p):
        wb = load_workbook(p)
        ws = wb["Sheet"] if "Sheet" in wb.sheetnames else wb.active
    else:
        wb = Workbook(); ws = wb.active; ws.title = "Sheet"
        ws.cell(1, 1, "지역"); ws.cell(1, 2, "입금자명"); ws.cell(1, 3, "대상거래처")
    idx = {}
    for r in range(2, ws.max_row + 1):
        loc, nm = ws.cell(r, 1).value, ws.cell(r, 2).value
        if loc and nm:
            idx[(str(loc).strip(), E.cname(nm))] = r
    nxt = ws.max_row + 1
    n = 0
    for loc, nm, tg in rows:
        loc = str(loc or "").strip(); nm = str(nm or "").strip(); tg = str(tg or "").strip()
        if not (loc and nm and tg): continue
        k = (loc, E.cname(nm))
        r = idx.get(k)
        if not r:
            r = nxt; nxt += 1; idx[k] = r
        ws.cell(r, 1, loc); ws.cell(r, 2, nm); ws.cell(r, 3, tg)
        n += 1
    wb.save(p)
    _zip_dir(work_dir, data_zip)
    return n
